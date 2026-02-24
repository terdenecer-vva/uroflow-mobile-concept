#!/usr/bin/env python3
"""
One-click daily QA runner for Uroflow Golden Dataset.

Example:
    python run_daily_qa.py --dataset_root /data/golden_ds --manifest /data/golden_ds/manifest.csv --out outputs

Outputs:
    outputs/YYYY-MM-DD/qa_record_level.csv
    outputs/YYYY-MM-DD/qa_summary.json
    outputs/YYYY-MM-DD/daily_qa_report.xlsx
    outputs/YYYY-MM-DD/daily_qa_report.pdf
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Dict, List

import numpy as np

from uroflow_qa_utils import (
    load_json,
    save_json,
    now_ymd,
    load_manifest,
    validate_manifest_rows,
    find_record_folder,
    parse_qref_csv,
    check_qref_sanity,
    integrate_flow,
    convert_audio_to_wav,
    detect_audio_onset,
    audio_proxy_q,
    resample_to_grid,
    pearson_corr,
    check_mp4_header,
    sha256_file,
)

DEFAULT_CONFIG_REL = "config/qa_config.json"


def write_csv(rows: List[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def generate_excel_report(summary: dict, record_rows: List[dict], out_xlsx: Path) -> None:
    """
    Creates a lightweight Excel report (Summary + Record_Level).
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Metric", "Value"])
    for k, v in summary.items():
        if isinstance(v, (dict, list)):
            continue
        ws.append([k, v])

    ws2 = wb.create_sheet("Record_Level")
    if record_rows:
        ws2.append(list(record_rows[0].keys()))
        for r in record_rows:
            ws2.append([r.get(k, "") for k in record_rows[0].keys()])

        # autosize
        for col_idx, col in enumerate(ws2.columns, start=1):
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def generate_pdf_report(summary: dict, record_rows: List[dict], out_pdf: Path, top_n: int = 15) -> None:
    """
    Generates a simple PDF daily report. If reportlab is not available, skips silently.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except Exception:
        return

    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(out_pdf), pagesize=A4)
    width, height = A4

    y = height - 20 * mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(20 * mm, y, "Uroflow Golden Dataset — Daily QA Report")
    y -= 8 * mm
    c.setFont("Helvetica", 10)
    c.drawString(20 * mm, y, f"Date: {summary.get('date', '')}")
    y -= 10 * mm

    # Summary block
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, "Summary")
    y -= 6 * mm
    c.setFont("Helvetica", 10)
    keys = [
        "n_records_manifest",
        "n_records_checked",
        "n_pass",
        "n_review",
        "n_fail",
    ]
    for k in keys:
        c.drawString(22 * mm, y, f"{k}: {summary.get(k, '')}")
        y -= 5 * mm

    y -= 4 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, "Top issues (FAIL/REVIEW)")
    y -= 6 * mm
    c.setFont("Helvetica", 9)

    # List top failing records
    bad = [r for r in record_rows if r.get("overall_status") in ("FAIL", "REVIEW")]
    bad = bad[:top_n]
    for r in bad:
        line = f"{r.get('record_id','')}: {r.get('overall_status','')} | {r.get('fail_codes','')}"
        if y < 20 * mm:
            c.showPage()
            y = height - 20 * mm
            c.setFont("Helvetica", 9)
        c.drawString(20 * mm, y, line[:120])
        y -= 5 * mm

    c.showPage()
    c.save()


def compute_dataset_checksums(dataset_root: Path, record_folders: List[Path], out_sha: Path) -> None:
    """
    SHA256 for all files in each record folder (recursive). Writes standard sha256sum format.
    """
    lines = []
    for rf in record_folders:
        for p in sorted(rf.rglob("*")):
            if p.is_file():
                h = sha256_file(p)
                rel = p.relative_to(dataset_root)
                lines.append(f"{h}  {rel.as_posix()}")
    out_sha.parent.mkdir(parents=True, exist_ok=True)
    out_sha.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset_root", type=str, required=True)
    ap.add_argument("--manifest", type=str, required=True)
    ap.add_argument("--out", type=str, default="outputs")
    ap.add_argument("--config", type=str, default="")
    ap.add_argument("--write_checksums", action="store_true")
    args = ap.parse_args()

    dataset_root = Path(args.dataset_root).expanduser().resolve()
    manifest_path = Path(args.manifest).expanduser().resolve()
    out_root = Path(args.out).expanduser().resolve()

    config_path = Path(args.config).expanduser().resolve() if args.config else (Path(__file__).parent.parent / DEFAULT_CONFIG_REL).resolve()
    config = load_json(config_path)

    date = now_ymd()
    out_dir = out_root / date
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load and validate manifest
    rows = load_manifest(manifest_path)
    rows, manifest_issues = validate_manifest_rows(rows, config)

    # Record-level checks
    record_results = []
    record_folders = []

    for r in rows:
        record_id = (r.get("record_id") or "").strip()
        if not record_id:
            continue

        rr = {
            "record_id": record_id,
            "overall_status": "PASS",
            "fail_codes": "",
            "record_folder": "",
            "missing_files": "",
            "qref_sanity": "",
            "Vvoid_ref_ml": r.get("Vvoid_ref_ml", ""),
            "Vvoid_int_ml": "",
            "Vvoid_delta_ml": "",
            "audio_present": "",
            "audio_onset_s": "",
            "audio_qref_corr": "",
            "video_present": "",
            "video_header_issue": "",
        }

        fail_codes = []
        # locate record folder
        rf = find_record_folder(dataset_root, record_id, config)
        if rf is None:
            rr["overall_status"] = "FAIL"
            fail_codes.append("RECORD_FOLDER_MISSING")
            record_results.append(rr)
            continue
        rr["record_folder"] = str(rf.relative_to(dataset_root))
        record_folders.append(rf)

        # required files
        missing = []
        for fn in config.get("required_record_files", []):
            if not (rf / fn).exists():
                missing.append(fn)
        rr["missing_files"] = ";".join(missing)
        if missing:
            rr["overall_status"] = "FAIL"
            fail_codes.append("MISSING_REQUIRED_FILES")

        # Q_ref checks
        qref_path = rf / "Q_ref.csv"
        t, q, v, qref_parse_issues = parse_qref_csv(qref_path, config)
        sanity_issues = check_qref_sanity(t, q) if t.size else ["Q_ref parse failed"]
        if qref_parse_issues:
            sanity_issues.extend(qref_parse_issues)
        rr["qref_sanity"] = ";".join(sanity_issues[:5])

        if sanity_issues:
            # treat time monotonic or too short as FAIL; other as REVIEW
            severe = any(("not strictly increasing" in s or "too short" in s or "missing required columns" in s) for s in sanity_issues)
            if severe and rr["overall_status"] != "FAIL":
                rr["overall_status"] = "FAIL"
                fail_codes.append("QREF_INVALID")
            elif rr["overall_status"] == "PASS":
                rr["overall_status"] = "REVIEW"
                fail_codes.append("QREF_SANITY_WARN")

        # Integral consistency vs Vvoid_ref_ml
        if t.size and q.size:
            Vint = integrate_flow(t, q)
            rr["Vvoid_int_ml"] = f"{Vint:.1f}"
            Vref = None
            try:
                Vref = float(str(r.get("Vvoid_ref_ml", "")).replace(",", "."))
            except Exception:
                Vref = None
            if Vref is not None and not np.isnan(Vint):
                delta = Vint - Vref
                rr["Vvoid_delta_ml"] = f"{delta:.1f}"
                abs_max = float(config.get("qref_integral_abs_ml_max", 10))
                pct_max = float(config.get("qref_integral_pct_max", 0.05))
                if abs(delta) > max(abs_max, pct_max * max(1.0, Vref)):
                    if rr["overall_status"] == "PASS":
                        rr["overall_status"] = "REVIEW"
                    fail_codes.append("QREF_INTEGRAL_MISMATCH")

        # Optional video integrity
        vid = rf / "roi_video.mp4"
        if vid.exists():
            rr["video_present"] = "yes"
            vh = check_mp4_header(vid)
            rr["video_header_issue"] = vh or ""
            if vh and rr["overall_status"] == "PASS":
                rr["overall_status"] = "REVIEW"
                fail_codes.append("VIDEO_INVALID")
        else:
            rr["video_present"] = "no"

        # Optional audio sync check
        audio_path = None
        if (rf / "audio.wav").exists():
            audio_path = rf / "audio.wav"
        elif (rf / "audio.m4a").exists():
            audio_path = rf / "audio.m4a"

        if audio_path is not None:
            rr["audio_present"] = "yes"
            try:
                wav = convert_audio_to_wav(audio_path)
                onset, debug, a_issues = detect_audio_onset(wav, config)
                if onset is not None:
                    rr["audio_onset_s"] = f"{onset:.2f}"
                    # correlate audio proxy with Q_ref
                    ta, qa, ap_issues = audio_proxy_q(wav, onset, target_hz=10.0)
                    if ta.size and t.size:
                        grid_t = t  # Q_ref is already at its grid
                        qa_rs = resample_to_grid(ta, qa, grid_t)
                        # normalize q_ref too
                        qn = q.copy()
                        if np.nanmax(qn) > 0:
                            qn = qn / np.nanmax(qn)
                        corr = pearson_corr(qa_rs, qn)
                        if corr is not None:
                            rr["audio_qref_corr"] = f"{corr:.2f}"
                            if corr < float(config.get("sync_corr_min", 0.25)):
                                if rr["overall_status"] == "PASS":
                                    rr["overall_status"] = "REVIEW"
                                fail_codes.append("AUDIO_QREF_LOW_CORR")
                        else:
                            if rr["overall_status"] == "PASS":
                                rr["overall_status"] = "REVIEW"
                            fail_codes.append("AUDIO_QREF_CORR_NA")
                    else:
                        if rr["overall_status"] == "PASS":
                            rr["overall_status"] = "REVIEW"
                        fail_codes.append("AUDIO_PROXY_FAIL")
                else:
                    rr["audio_onset_s"] = ""
                    if rr["overall_status"] == "PASS":
                        rr["overall_status"] = "REVIEW"
                    fail_codes.append("AUDIO_ONSET_NOT_FOUND")
            except Exception as e:
                rr["audio_present"] = "error"
                rr["audio_onset_s"] = ""
                if rr["overall_status"] == "PASS":
                    rr["overall_status"] = "REVIEW"
                fail_codes.append(f"AUDIO_ERROR:{type(e).__name__}")
        else:
            rr["audio_present"] = "no"

        rr["fail_codes"] = ";".join(fail_codes)
        record_results.append(rr)

    # Summarize
    n_pass = sum(1 for r in record_results if r["overall_status"] == "PASS")
    n_review = sum(1 for r in record_results if r["overall_status"] == "REVIEW")
    n_fail = sum(1 for r in record_results if r["overall_status"] == "FAIL")

    summary = {
        "date": date,
        "dataset_root": str(dataset_root),
        "manifest": str(manifest_path),
        "config": str(config_path),
        "n_records_manifest": len(rows),
        "n_records_checked": len(record_results),
        "n_pass": n_pass,
        "n_review": n_review,
        "n_fail": n_fail,
        "manifest_issue_count": len(manifest_issues),
    }

    # Save outputs
    write_csv(record_results, out_dir / "qa_record_level.csv")
    save_json(summary, out_dir / "qa_summary.json")
    if manifest_issues:
        write_csv(manifest_issues, out_dir / "qa_manifest_issues.csv")

    # checksums
    if args.write_checksums:
        compute_dataset_checksums(dataset_root, record_folders, out_dir / "checksums.sha256")

    # reports
    try:
        generate_excel_report(summary, record_results, out_dir / "daily_qa_report.xlsx")
    except Exception:
        pass
    try:
        top_n = int(config.get("daily_report_top_n_failures", 15))
        generate_pdf_report(summary, record_results, out_dir / "daily_qa_report.pdf", top_n=top_n)
    except Exception:
        pass

    print(f"[OK] QA finished. Outputs: {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
