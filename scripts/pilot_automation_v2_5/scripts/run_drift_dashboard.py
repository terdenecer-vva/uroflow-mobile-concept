#!/usr/bin/env python3
"""
Drift dashboard generator (offline)

Generates stratified performance summaries from the golden dataset:
- by site_id / toilet_id / iphone_model / noise_level / posture / sex
- flags drift vs overall baseline

Input:
- tfl_record_level.csv (recommended) OR dataset_root+manifest (will compute quickly)

Output:
- drift_dashboard.xlsx
- drift_summary.json
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd

from uroflow_qa_utils import load_manifest
from run_tfl_from_golden_dataset import ba_stats


def save_json(obj: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)


def compute_group_metrics(df: pd.DataFrame, group_col: str, overall: dict, cfg: dict) -> pd.DataFrame:
    crit = cfg["criteria"]
    min_n = int(crit.get("min_records_per_group", 10))
    drift_mult = float(crit.get("drift_mae_multiplier_max", 1.5))

    rows = []
    for g, dfg in df.groupby(group_col):
        n_total = int(dfg.shape[0])
        dvalid = dfg[dfg["valid_for_primary"] == "Y"]
        n_valid = int(dvalid.shape[0])
        valid_rate = (n_valid / n_total) if n_total else math.nan

        if n_valid >= 3:
            st_qmax = ba_stats(dvalid["ref_Qmax_ml_s"].to_numpy(float), dvalid["app_Qmax_ml_s"].to_numpy(float))
            st_v = ba_stats(dvalid["ref_Vvoid_ml"].to_numpy(float), dvalid["app_Vvoid_ml"].to_numpy(float))
        else:
            st_qmax = {"mae": math.nan, "bias": math.nan, "mape": math.nan, "n": n_valid}
            st_v = {"mae": math.nan, "bias": math.nan, "mape": math.nan, "n": n_valid}

        # drift flags
        drift_flags = []
        if n_total >= min_n and math.isfinite(overall.get("Qmax_mae", math.nan)) and math.isfinite(st_qmax.get("mae", math.nan)):
            if st_qmax["mae"] > overall["Qmax_mae"] * drift_mult:
                drift_flags.append("Qmax_MAE_DRIFT")
        if n_total >= min_n and math.isfinite(overall.get("Vvoid_mape", math.nan)) and math.isfinite(st_v.get("mape", math.nan)):
            if st_v["mape"] > overall["Vvoid_mape"] * drift_mult:
                drift_flags.append("Vvoid_MAPE_DRIFT")

        rows.append({
            group_col: g,
            "n_total": n_total,
            "n_valid": n_valid,
            "valid_rate": valid_rate,
            "Qmax_mae": st_qmax.get("mae", math.nan),
            "Qmax_bias": st_qmax.get("bias", math.nan),
            "Vvoid_mape": st_v.get("mape", math.nan),
            "Vvoid_mae": st_v.get("mae", math.nan),
            "flags": ";".join(drift_flags) if drift_flags else "",
        })

    out = pd.DataFrame(rows)
    out = out.sort_values(by=["flags", "valid_rate"], ascending=[False, True])
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--tfl_csv", default="", help="Path to tfl_record_level.csv")
    ap.add_argument("--out_dir", default=str(Path(__file__).resolve().parents[1] / "outputs" / "drift"), help="Output directory")
    ap.add_argument("--g1_config", default=str(Path(__file__).resolve().parents[1] / "config" / "g1_acceptance_config.json"))
    args = ap.parse_args()

    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if not args.tfl_csv:
        raise SystemExit("Please provide --tfl_csv (output of run_tfl_from_golden_dataset.py)")

    df = pd.read_csv(args.tfl_csv)
    # overall baseline computed on valid records
    dv = df[df["valid_for_primary"] == "Y"]
    overall = {}
    if dv.shape[0] >= 3:
        st_qmax = ba_stats(dv["ref_Qmax_ml_s"].to_numpy(float), dv["app_Qmax_ml_s"].to_numpy(float))
        st_v = ba_stats(dv["ref_Vvoid_ml"].to_numpy(float), dv["app_Vvoid_ml"].to_numpy(float))
        overall["Qmax_mae"] = st_qmax["mae"]
        overall["Vvoid_mape"] = st_v["mape"]
    else:
        overall["Qmax_mae"] = math.nan
        overall["Vvoid_mape"] = math.nan

    cfg = json.loads(Path(args.g1_config).read_text(encoding="utf-8"))
    groupings = ["site_id","toilet_id","iphone_model","noise_level","sex","posture"]
    wb_path = out_dir / "drift_dashboard.xlsx"

    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = openpyxl.Workbook()
    # remove default
    wb.remove(wb.active)

    summary = {"overall": overall, "groups": {}}

    for gc in groupings:
        if gc not in df.columns:
            continue
        tab = compute_group_metrics(df, gc, overall, cfg)
        summary["groups"][gc] = {
            "n_groups": int(tab.shape[0]),
        }
        ws = wb.create_sheet(title=gc[:31])
        for r in dataframe_to_rows(tab, index=False, header=True):
            ws.append(r)

    ws0 = wb.create_sheet(title="README")
    ws0["A1"] = "Drift dashboard (auto-generated). Review groups with flags and low valid_rate."
    ws0["A3"] = f"Overall baseline: Qmax MAE={overall.get('Qmax_mae')} | Vvoid MAPE={overall.get('Vvoid_mape')}"

    wb.save(wb_path)
    save_json(summary, out_dir / "drift_summary.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
