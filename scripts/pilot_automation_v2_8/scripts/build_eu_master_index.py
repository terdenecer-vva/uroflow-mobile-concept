#!/usr/bin/env python3
"""Build an executed EU MDR master index with SHA256 + Evidence_ID mapping.

Inputs:
- EU Annex II/III submission folder index (xlsx)
- EU GSPR checklist (xlsx) for additional references
- Pilot Readiness Evidence Manifest (xlsx) for Evidence_ID mapping (optional)

Output:
- An Excel workbook with:
  - Master_Index: de-duplicated file list (Annex + GSPR) with existence + SHA256
  - Missing: missing files with source/where referenced
  - Summary

Usage:
  python build_eu_master_index.py --build_root <Submission_Build> \
    --annex_index 06_EU_MDR/Annex_II_III_Submission_Folder/EU_MDR_AnnexII_III_Submission_Folder_Index_v2.2.xlsx \
    --gspr 06_EU_MDR/Uroflow_EU_MDR_GSPR_Checklist_AnnexI_v1.2_EXECUTED_AUTO.xlsx \
    --evidence_manifest 05_Clinical/Uroflow_P0_Pilot_Readiness_Evidence_Manifest_v1.0.xlsx \
    --out 06_EU_MDR/Uroflow_EU_MDR_AnnexII_III_Master_Index_v1.0_EXECUTED.xlsx
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

from uroflow_qa_utils import sha256_file


def _load_index_rows(index_path: Path) -> List[dict]:
    wb = openpyxl.load_workbook(index_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    header = [str(h).strip() if h is not None else "" for h in header]

    def get(row, name, default=""):
        if name not in header:
            return default
        return ws.cell(row, header.index(name) + 1).value

    rows = []
    for r in range(2, ws.max_row + 1):
        if all((ws.cell(r, c).value is None or str(ws.cell(r, c).value).strip() == "") for c in range(1, min(ws.max_column, 6) + 1)):
            continue
        rows.append({
            "section": str(get(r, "Section", "")).strip(),
            "file_name": str(get(r, "File name", "")).strip(),
            "rel_path": str(get(r, "Relative path", "")).strip(),
            "status": str(get(r, "Status", "")).strip(),
            "notes": str(get(r, "Notes", "")).strip(),
        })
    return rows


def _load_evidence_map(evidence_manifest: Path) -> Dict[str, str]:
    if not evidence_manifest.exists():
        return {}
    wb = openpyxl.load_workbook(evidence_manifest, data_only=True)
    sheet = "Manifest" if "Manifest" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    header = [c.value for c in ws[1]]
    header = [str(h).strip() if h is not None else "" for h in header]

    def idx(name: str) -> int:
        return header.index(name) + 1 if name in header else -1

    col_eid = idx("Evidence_ID")
    col_expected = idx("Expected file name")
    if col_eid < 0 or col_expected < 0:
        return {}

    mp: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        eid = ws.cell(r, col_eid).value
        exp = ws.cell(r, col_expected).value
        if eid is None or exp is None:
            continue
        eid = str(eid).strip()
        exp = str(exp).strip()
        if eid and exp:
            mp[exp] = eid
    return mp


def _parse_gspr_refs(gspr_path: Path) -> List[Tuple[str, str]]:
    """Return list of (gspr_id, rel_path) from GSPR sheet."""
    wb = openpyxl.load_workbook(gspr_path, data_only=True)
    if "GSPR_Checklist" not in wb.sheetnames:
        return []
    ws = wb["GSPR_Checklist"]

    # Find columns by header
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header = [str(h).strip() if h is not None else "" for h in header]

    def col(name: str, default: int) -> int:
        return header.index(name) + 1 if name in header else default

    col_gspr_id = col("GSPR ID (Annex I)", 1)
    col_refs = col("Build file references (relative paths)", 9)

    pairs: List[Tuple[str, str]] = []
    for r in range(2, ws.max_row + 1):
        gid = ws.cell(r, col_gspr_id).value
        gid = str(gid).strip() if gid is not None else ""
        refs = ws.cell(r, col_refs).value
        if refs is None:
            continue
        for rel in str(refs).splitlines():
            rel = rel.strip()
            if not rel:
                continue
            pairs.append((gid, rel))
    return pairs


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--build_root", required=True)
    ap.add_argument("--annex_index", required=True)
    ap.add_argument("--gspr", required=True)
    ap.add_argument("--evidence_manifest", default="")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    build_root = Path(args.build_root).expanduser().resolve()

    annex_index = Path(args.annex_index)
    if not annex_index.is_absolute():
        annex_index = (build_root / annex_index).resolve()

    gspr_path = Path(args.gspr)
    if not gspr_path.is_absolute():
        gspr_path = (build_root / gspr_path).resolve()

    evidence_manifest = Path(args.evidence_manifest) if args.evidence_manifest else Path()
    if evidence_manifest and not evidence_manifest.is_absolute():
        evidence_manifest = (build_root / evidence_manifest).resolve()

    out_path = Path(args.out)
    if not out_path.is_absolute():
        out_path = (build_root / out_path).resolve()

    if not annex_index.exists():
        raise FileNotFoundError(f"Annex index not found: {annex_index}")
    if not gspr_path.exists():
        raise FileNotFoundError(f"GSPR file not found: {gspr_path}")

    evidence_map = _load_evidence_map(evidence_manifest) if evidence_manifest else {}

    # Collect items
    items: Dict[str, dict] = {}  # rel_path -> record

    annex_rows = _load_index_rows(annex_index)
    for r in annex_rows:
        rel = r["rel_path"]
        if not rel:
            continue
        items.setdefault(rel, {
            "rel_path": rel,
            "file_name": r["file_name"] or Path(rel).name,
            "sources": set(),
            "sections": set(),
            "gspr_ids": set(),
            "annex_status": r.get("status", ""),
        })
        items[rel]["sources"].add("ANNEX_II_III")
        if r.get("section"):
            items[rel]["sections"].add(r["section"])

    gspr_pairs = _parse_gspr_refs(gspr_path)
    for gid, rel in gspr_pairs:
        if not rel:
            continue
        items.setdefault(rel, {
            "rel_path": rel,
            "file_name": Path(rel).name,
            "sources": set(),
            "sections": set(),
            "gspr_ids": set(),
            "annex_status": "",
        })
        items[rel]["sources"].add("GSPR")
        if gid:
            items[rel]["gspr_ids"].add(gid)

    # Enrich with existence + SHA + evidence id
    missing_rows = []
    for rel, rec in items.items():
        p = (build_root / rel).resolve()
        rec["exists"] = "Y" if p.exists() and p.is_file() else "N"
        rec["sha256"] = sha256_file(p) if rec["exists"] == "Y" else ""
        rec["sha256_short"] = rec["sha256"][:12] if rec["sha256"] else ""
        eid = evidence_map.get(Path(rel).name)
        rec["evidence_id"] = eid or ""
        if rec["exists"] == "N":
            missing_rows.append({
                "rel_path": rel,
                "file_name": rec["file_name"],
                "sources": ",".join(sorted(rec["sources"])),
                "sections": ",".join(sorted(rec["sections"])),
                "gspr_ids": ",".join(sorted(rec["gspr_ids"])),
            })

    # Write workbook
    wb_out = openpyxl.Workbook()
    ws_m = wb_out.active
    ws_m.title = "Master_Index"

    headers = [
        "rel_path",
        "file_name",
        "sources",
        "annex_sections",
        "gspr_ids",
        "annex_status",
        "exists",
        "sha256",
        "sha256_short",
        "evidence_id",
    ]
    ws_m.append(headers)

    for rel in sorted(items.keys()):
        rec = items[rel]
        ws_m.append([
            rec["rel_path"],
            rec["file_name"],
            ",".join(sorted(rec["sources"])),
            ",".join(sorted(rec["sections"])),
            ",".join(sorted(rec["gspr_ids"])),
            rec.get("annex_status", ""),
            rec.get("exists", ""),
            rec.get("sha256", ""),
            rec.get("sha256_short", ""),
            rec.get("evidence_id", ""),
        ])

    # Missing sheet
    ws_missing = wb_out.create_sheet("Missing")
    ws_missing.append(["rel_path", "file_name", "sources", "annex_sections", "gspr_ids"])
    for mr in sorted(missing_rows, key=lambda x: x["rel_path"]):
        ws_missing.append([mr["rel_path"], mr["file_name"], mr["sources"], mr["sections"], mr["gspr_ids"]])

    # Summary
    ws_sum = wb_out.create_sheet("Summary")
    ws_sum["A1"].value = "Build root"
    ws_sum["B1"].value = str(build_root)
    ws_sum["A2"].value = "Annex index"
    ws_sum["B2"].value = str(annex_index)
    ws_sum["A3"].value = "GSPR"
    ws_sum["B3"].value = str(gspr_path)
    ws_sum["A4"].value = "Evidence manifest"
    ws_sum["B4"].value = str(evidence_manifest) if evidence_manifest else "(none)"

    total = len(items)
    present = sum(1 for r in items.values() if r.get("exists") == "Y")
    ws_sum["A6"].value = "Total files (unique)"
    ws_sum["B6"].value = total
    ws_sum["A7"].value = "Present"
    ws_sum["B7"].value = present
    ws_sum["A8"].value = "Missing"
    ws_sum["B8"].value = total - present

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out_path)
    print(f"[OK] Wrote: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
