#!/usr/bin/env python3
"""Autofill EU MDR GSPR checklist with *evidence file links* based on Evidence_ID.

Goal
- Extend the file-level GSPR execution with evidence-bundle traceability:
  Evidence_ID -> expected filename -> discovered file path(s) -> SHA256

Inputs
- build_root: root of Submission_Build
- gspr_in: executed GSPR XLSX (e.g., v1.2_EXECUTED_AUTO)
- evidence_manifest: Pilot Readiness Evidence Manifest (Evidence_ID + Expected file name)

Search strategy
- By default, searches these folders (relative to build_root):
  - 12_Evidence_Bundles/Evidence_Files
  - 10_Pilot_Automation/outputs
  - 05_Clinical
  - 01_Product_QMS

Output
- Writes a new XLSX with extra columns:
  - Autofill: Evidence files found (relative paths)
  - Autofill: Evidence SHA256 short (basename:hash12)
  - Autofill: Evidence availability

Usage
  python autofill_gspr_evidence_links.py \
    --build_root <Submission_Build_vX.X> \
    --gspr_in 06_EU_MDR/Uroflow_EU_MDR_GSPR_Checklist_AnnexI_v1.2_EXECUTED_AUTO.xlsx \
    --gspr_out 06_EU_MDR/Uroflow_EU_MDR_GSPR_Checklist_AnnexI_v1.3_EXECUTED_EVIDENCE_AUTO.xlsx \
    --evidence_manifest 05_Clinical/Uroflow_P0_Pilot_Readiness_Evidence_Manifest_v1.0.xlsx

Notes
- This script is intentionally conservative: it does not change any original GSPR content.
- Evidence discovery is filename-based to keep it robust across folder layouts.
"""

from __future__ import annotations

import argparse
import os
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

from uroflow_qa_utils import sha256_file


SKIP_DIR_NAMES = {
    "Archive",
    ".git",
    "__pycache__",
    "records",  # may contain large raw data
    "raw",
    "videos",
    "audio",
}


def _parse_multiline(cell_value) -> List[str]:
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.splitlines()]
    return [p for p in parts if p]


def _load_evidence_manifest(path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Returns:
    - eid_to_expected_filename
    - expected_filename_to_eid
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Manifest"] if "Manifest" in wb.sheetnames else wb[wb.sheetnames[0]]

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header = [str(h).strip() if h is not None else "" for h in header]

    def col(name: str) -> int:
        try:
            return header.index(name) + 1
        except ValueError:
            return -1

    c_eid = col("Evidence_ID")
    c_exp = col("Expected file name")
    if c_eid < 0 or c_exp < 0:
        raise ValueError("Evidence manifest must contain 'Evidence_ID' and 'Expected file name' columns")

    eid_to_fn: Dict[str, str] = {}
    fn_to_eid: Dict[str, str] = {}

    for r in range(2, ws.max_row + 1):
        eid = ws.cell(r, c_eid).value
        fn = ws.cell(r, c_exp).value
        if eid is None or fn is None:
            continue
        eid = str(eid).strip()
        fn = str(fn).strip()
        if not eid or not fn:
            continue
        eid_to_fn[eid] = fn
        # If duplicates exist, keep the first mapping
        fn_to_eid.setdefault(fn, eid)

    return eid_to_fn, fn_to_eid


def _walk_index_files(root: Path) -> Dict[str, List[Path]]:
    """Return filename -> list of relative paths (to build_root)."""
    index: Dict[str, List[Path]] = defaultdict(list)

    for dirpath, dirnames, filenames in os.walk(root):
        # prune dirs
        dirnames[:] = [d for d in dirnames if d not in SKIP_DIR_NAMES]

        for fn in filenames:
            if not fn or fn.startswith("~$"):
                continue
            p = Path(dirpath) / fn
            index[fn].append(p)

    return index


def _ensure_columns(ws, headers: List[str]) -> List[int]:
    existing = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    existing = [str(v).strip() if v is not None else "" for v in existing]

    cols = []
    next_col = ws.max_column + 1
    for h in headers:
        if h in existing:
            cols.append(existing.index(h) + 1)
        else:
            ws.cell(1, next_col).value = h
            cols.append(next_col)
            next_col += 1
    return cols


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--build_root", required=True)
    ap.add_argument("--gspr_in", required=True)
    ap.add_argument("--gspr_out", required=True)
    ap.add_argument("--evidence_manifest", required=True)
    ap.add_argument(
        "--search_roots",
        default="12_Evidence_Bundles/Evidence_Files;10_Pilot_Automation/outputs;05_Clinical;01_Product_QMS",
        help="Semicolon-separated search roots (relative to build_root)",
    )
    args = ap.parse_args()

    build_root = Path(args.build_root).expanduser().resolve()

    gspr_in = Path(args.gspr_in)
    if not gspr_in.is_absolute():
        gspr_in = (build_root / gspr_in).resolve()

    gspr_out = Path(args.gspr_out)
    if not gspr_out.is_absolute():
        gspr_out = (build_root / gspr_out).resolve()

    evidence_manifest = Path(args.evidence_manifest)
    if not evidence_manifest.is_absolute():
        evidence_manifest = (build_root / evidence_manifest).resolve()

    if not gspr_in.exists():
        raise FileNotFoundError(f"GSPR input not found: {gspr_in}")
    if not evidence_manifest.exists():
        raise FileNotFoundError(f"Evidence manifest not found: {evidence_manifest}")

    eid_to_fn, _ = _load_evidence_manifest(evidence_manifest)

    # Build file index across selected roots
    search_roots = []
    for part in str(args.search_roots).split(";"):
        part = part.strip()
        if not part:
            continue
        p = (build_root / part).resolve()
        if p.exists() and p.is_dir():
            search_roots.append(p)

    file_index: Dict[str, List[Path]] = defaultdict(list)
    for sr in search_roots:
        idx = _walk_index_files(sr)
        for fn, paths in idx.items():
            file_index[fn].extend(paths)

    wb = openpyxl.load_workbook(gspr_in)
    if "GSPR_Checklist" not in wb.sheetnames:
        raise ValueError("Sheet 'GSPR_Checklist' not found")

    ws = wb["GSPR_Checklist"]

    # Find column indices by header names
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    def find_col(name: str) -> int:
        try:
            return headers.index(name) + 1
        except ValueError:
            return -1

    col_eids = find_col("Autofill: Evidence_ID matches")
    # fallback: the file-level autofill uses this column name; if missing, we can't proceed reliably
    if col_eids < 0:
        raise ValueError("Column 'Autofill: Evidence_ID matches' not found. Run file-level GSPR autofill first.")

    # Add new columns
    new_headers = [
        "Autofill: Evidence files found (relative paths)",
        "Autofill: Evidence SHA256 short (basename:hash12)",
        "Autofill: Evidence availability (ALL_PRESENT/PARTIAL/ALL_MISSING/NO_EVIDENCE_IDS)",
    ]
    col_paths, col_sha, col_avail = _ensure_columns(ws, new_headers)

    missing_evidence_counter = Counter()
    missing_file_counter = Counter()

    total_rows = 0
    rows_with_eids = 0
    rows_all_present = 0

    for r in range(2, ws.max_row + 1):
        total_rows += 1
        eids = _parse_multiline(ws.cell(r, col_eids).value)
        if not eids:
            ws.cell(r, col_paths).value = ""
            ws.cell(r, col_sha).value = ""
            ws.cell(r, col_avail).value = "NO_EVIDENCE_IDS"
            continue

        rows_with_eids += 1
        found_paths: List[str] = []
        sha_entries: List[str] = []
        found = 0

        for eid in eids:
            expected_fn = eid_to_fn.get(eid)
            if not expected_fn:
                missing_evidence_counter[eid] += 1
                continue

            candidates = file_index.get(expected_fn, [])
            if not candidates:
                missing_file_counter[expected_fn] += 1
                continue

            # take all candidates, but de-dup
            uniq = []
            seen = set()
            for p in candidates:
                rel = str(p.relative_to(build_root)) if p.is_absolute() else str(p)
                if rel in seen:
                    continue
                seen.add(rel)
                uniq.append((p, rel))

            for p, rel in uniq:
                found_paths.append(rel)
                try:
                    h = sha256_file(p)
                    sha_entries.append(f"{Path(rel).name}:{h[:12]}")
                except Exception:
                    sha_entries.append(f"{Path(rel).name}:sha_error")

            found += 1

        # Write outputs
        ws.cell(r, col_paths).value = "\n".join(found_paths)
        ws.cell(r, col_sha).value = "\n".join(sha_entries)

        if found == len(eids):
            avail = "ALL_PRESENT"
            rows_all_present += 1
        elif found == 0:
            avail = "ALL_MISSING"
        else:
            avail = "PARTIAL"
        ws.cell(r, col_avail).value = avail

    # Summary sheet
    if "Autofill_Evidence_Summary" in wb.sheetnames:
        del wb["Autofill_Evidence_Summary"]
    ws_sum = wb.create_sheet("Autofill_Evidence_Summary")

    ws_sum["A1"].value = "Build root"
    ws_sum["B1"].value = str(build_root)
    ws_sum["A2"].value = "GSPR input"
    ws_sum["B2"].value = str(gspr_in)
    ws_sum["A3"].value = "Evidence manifest"
    ws_sum["B3"].value = str(evidence_manifest)
    ws_sum["A4"].value = "Search roots"
    ws_sum["B4"].value = " ; ".join([str(p.relative_to(build_root)) for p in search_roots])

    ws_sum["A6"].value = "Total rows"
    ws_sum["B6"].value = total_rows
    ws_sum["A7"].value = "Rows with Evidence_IDs"
    ws_sum["B7"].value = rows_with_eids
    ws_sum["A8"].value = "Rows ALL_PRESENT"
    ws_sum["B8"].value = rows_all_present

    ws_sum["A10"].value = "Missing Evidence_ID (not in manifest)"
    ws_sum["B10"].value = "Count"
    row = 11
    for k, v in missing_evidence_counter.most_common():
        ws_sum.cell(row, 1).value = k
        ws_sum.cell(row, 2).value = v
        row += 1

    row += 1
    ws_sum.cell(row, 1).value = "Missing expected filename (not found on disk)"
    ws_sum.cell(row, 2).value = "Count"
    row += 1
    for k, v in missing_file_counter.most_common():
        ws_sum.cell(row, 1).value = k
        ws_sum.cell(row, 2).value = v
        row += 1

    gspr_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(gspr_out)
    print(f"[OK] Wrote: {gspr_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
