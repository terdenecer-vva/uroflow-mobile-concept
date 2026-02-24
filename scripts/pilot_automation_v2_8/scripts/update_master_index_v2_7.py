#!/usr/bin/env python3
"""Update the Master Submission Index workbook for v2.7.

This is a small helper used during build packaging.
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--index_path", required=True)
    args = ap.parse_args()

    index_path = Path(args.index_path).expanduser().resolve()
    wb = openpyxl.load_workbook(index_path)

    if "File_Index" not in wb.sheetnames:
        raise SystemExit("File_Index sheet not found")

    ws = wb["File_Index"]

    entries = [
        ["00_README_and_Indexes","Multi","00_README_and_Indexes/00_README_Submission_Build_v2.7.txt","00_README_Submission_Build_v2.7.txt","TXT","2.7"],
        ["06_EU_MDR","EU","06_EU_MDR/Uroflow_EU_MDR_GSPR_Checklist_AnnexI_v1.2_EXECUTED_AUTO.xlsx","Uroflow_EU_MDR_GSPR_Checklist_AnnexI_v1.2_EXECUTED_AUTO.xlsx","XLSX","2.7"],
        ["06_EU_MDR","EU","06_EU_MDR/Uroflow_EU_MDR_AnnexII_III_Master_Index_v1.0_EXECUTED.xlsx","Uroflow_EU_MDR_AnnexII_III_Master_Index_v1.0_EXECUTED.xlsx","XLSX","2.7"],
        ["10.7","EU MDR automation – GSPR autofill (exists+SHA+Evidence_ID)","autofill_gspr_executed.py","10_Pilot_Automation/scripts/","New","Writes Annex I GSPR checklist as executed (file presence + hashes)"],
        ["10.8","EU MDR automation – Annex II/III master index (executed)","build_eu_master_index.py","10_Pilot_Automation/scripts/","New","De-duplicates Annex index + GSPR references; adds SHA256 + Evidence_ID mapping"],
        ["10.9","Automation – Pilot-freeze submission tree builder (EU+US)","build_pilotfreeze_submission_tree.py","10_Pilot_Automation/scripts/","New","Builds clean submission tree w/o archives/duplicates; generates checksums+manifest"],
        ["10.10","One-click runner: GSPR autofill","run_gspr_autofill_oneclick.sh / .bat","10_Pilot_Automation/","New","Wrapper around autofill_gspr_executed.py"],
        ["10.11","One-click runner: EU master index","run_eu_master_index_oneclick.sh / .bat","10_Pilot_Automation/","New","Wrapper around build_eu_master_index.py"],
        ["10.12","One-click runner: pilot-freeze tree","run_pilotfreeze_tree_oneclick.sh / .bat","10_Pilot_Automation/","New","Wrapper around build_pilotfreeze_submission_tree.py"],
        ["10_Pilot_Automation","Multi","10_Pilot_Automation/config/pilotfreeze_extra_includes.txt","config/pilotfreeze_extra_includes.txt","TXT","2.7"],
    ]

    for row in entries:
        ws.append(row)

    if "v2.7_additions" in wb.sheetnames:
        del wb["v2.7_additions"]

    ws_add = wb.create_sheet("v2.7_additions")
    ws_add["A1"].value = "Submission Build v2.7 – additions"
    ws_add["A3"].value = "1) GSPR executed autofill: adds existence/SHA256/Evidence_ID mapping for each GSPR row."
    ws_add["A4"].value = "2) EU Annex II/III master index executed: de-duplicated file list with hashes + missing report."
    ws_add["A5"].value = "3) Pilot-freeze submission tree builder: clean EU+US tree (no archives/duplicates) + checksums."
    ws_add["A7"].value = "Key outputs generated in this build:"
    ws_add["A8"].value = " - 06_EU_MDR/Uroflow_EU_MDR_GSPR_Checklist_AnnexI_v1.2_EXECUTED_AUTO.xlsx"
    ws_add["A9"].value = " - 06_EU_MDR/Uroflow_EU_MDR_AnnexII_III_Master_Index_v1.0_EXECUTED.xlsx"

    wb.save(index_path)
    print(f"[OK] Updated: {index_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
