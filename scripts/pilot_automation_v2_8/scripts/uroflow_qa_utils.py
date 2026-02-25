#!/usr/bin/env python3
"""
Uroflow QA Utilities (offline)

Keep dependencies lightweight:
- numpy, pandas (optional), openpyxl, reportlab (optional), ffmpeg (optional)
"""

from __future__ import annotations

import csv
import datetime as _dt
import json
import math
import os
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np


def load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(obj: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)


def now_ymd() -> str:
    return _dt.datetime.now().strftime("%Y-%m-%d")


def safe_float(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    # accept comma decimal
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def safe_int(x) -> Optional[int]:
    f = safe_float(x)
    if f is None:
        return None
    return int(round(f))


def load_manifest(manifest_path: Path) -> List[dict]:
    """
    Supports CSV and XLSX.
    For XLSX, reads sheet 'Manifest_Template' if present; else first sheet.
    """
    if not manifest_path.exists():
        raise FileNotFoundError(f"Manifest not found: {manifest_path}")

    suffix = manifest_path.suffix.lower()
    if suffix == ".csv":
        with open(manifest_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = []
            for row in reader:
                # Normalize keys
                rows.append({(k or "").strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()})
            return rows

    if suffix in [".xlsx", ".xlsm"]:
        try:
            import pandas as pd  # optional
            xls = pd.ExcelFile(manifest_path)
            sheet = "Manifest_Template" if "Manifest_Template" in xls.sheet_names else xls.sheet_names[0]
            df = pd.read_excel(manifest_path, sheet_name=sheet, dtype=str)
            df = df.fillna("")
            rows = df.to_dict(orient="records")
            # strip keys/values
            cleaned = []
            for r in rows:
                cleaned.append({str(k).strip(): (str(v).strip() if v is not None else "") for k, v in r.items()})
            return cleaned
        except Exception:
            # fallback to openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(manifest_path, data_only=True)
            sheet = "Manifest_Template" if "Manifest_Template" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            header = [str(h).strip() if h is not None else "" for h in header]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                r = {}
                for k, v in zip(header, row):
                    if not k:
                        continue
                    r[k] = str(v).strip() if v is not None else ""
                rows.append(r)
            return rows

    raise ValueError(f"Unsupported manifest type: {manifest_path.suffix}")


def validate_manifest_rows(rows: List[dict], config: dict) -> Tuple[List[dict], List[dict]]:
    """
    Returns: (clean_rows, issues)
    issues: list of dicts with keys: record_id, issue_code, field, message, severity
    """
    expected_cols = config.get("manifest_expected_columns", [])
    required_nonempty = set(config.get("manifest_required_nonempty_fields", []))
    codelists = config.get("codelists", {})

    issues = []
    clean = []
    seen_record_ids: set[str] = set()
    seen_sync_ids: dict[str, str] = {}

    # Validate columns presence
    if rows:
        present = set(rows[0].keys())
        missing_cols = [c for c in expected_cols if c not in present]
        if missing_cols:
            issues.append({
                "record_id": "",
                "issue_code": "MANIFEST_MISSING_COLUMNS",
                "field": "",
                "message": f"Missing columns: {', '.join(missing_cols)}",
                "severity": "FAIL",
            })

    for r in rows:
        record_id = (r.get("record_id") or "").strip()
        if not record_id:
            issues.append({
                "record_id": "",
                "issue_code": "RECORD_ID_MISSING",
                "field": "record_id",
                "message": "record_id is empty",
                "severity": "FAIL",
            })
            continue

        if record_id in seen_record_ids:
            issues.append({
                "record_id": record_id,
                "issue_code": "DUPLICATE_RECORD_ID",
                "field": "record_id",
                "message": f"Duplicate record_id in manifest: {record_id}",
                "severity": "FAIL",
            })
        else:
            seen_record_ids.add(record_id)

        sync_id = (r.get("sync_id") or "").strip()
        if sync_id:
            previous_record = seen_sync_ids.get(sync_id)
            if previous_record is not None and previous_record != record_id:
                issues.append({
                    "record_id": record_id,
                    "issue_code": "SYNC_ID_DUPLICATE",
                    "field": "sync_id",
                    "message": (
                        f"sync_id '{sync_id}' is reused by multiple record_id values: "
                        f"{previous_record}, {record_id}"
                    ),
                    "severity": "FAIL",
                })
            else:
                seen_sync_ids[sync_id] = record_id

        # required non-empty fields
        for f in required_nonempty:
            val = (r.get(f) or "").strip()
            if val == "":
                issues.append({
                    "record_id": record_id,
                    "issue_code": "REQUIRED_FIELD_EMPTY",
                    "field": f,
                    "message": f"Required field empty: {f}",
                    "severity": "FAIL",
                })

        # codelists
        for field, allowed in codelists.items():
            val = (r.get(field) or "").strip()
            if val == "":
                continue
            if field == "noise_source":
                # allow multiple values separated by ';'
                parts = [p.strip() for p in val.split(";") if p.strip()]
                bad = [p for p in parts if p not in allowed]
                if bad:
                    issues.append({
                        "record_id": record_id,
                        "issue_code": "CODELIST_VIOLATION",
                        "field": field,
                        "message": f"Invalid values: {bad}; allowed: {allowed}",
                        "severity": "REVIEW",
                    })
            else:
                if val not in allowed:
                    issues.append({
                        "record_id": record_id,
                        "issue_code": "CODELIST_VIOLATION",
                        "field": field,
                        "message": f"Invalid value '{val}'; allowed: {allowed}",
                        "severity": "REVIEW",
                    })

        clean.append(r)

    return clean, issues


def find_record_folder(dataset_root: Path, record_id: str, config: dict) -> Optional[Path]:
    candidates = config.get("record_folder_candidates", ["records/{record_id}", "{record_id}"])
    for pat in candidates:
        p = dataset_root / pat.format(record_id=record_id)
        if p.exists() and p.is_dir():
            return p
    return None


def parse_qref_csv(qref_path: Path, config: dict) -> Tuple[np.ndarray, np.ndarray, Optional[np.ndarray], List[str]]:
    """
    Reads Q_ref.csv and returns t_s, Q_ml_s, V_ml(optional) and list of parse issues.
    """
    issues = []
    if not qref_path.exists():
        return np.array([]), np.array([]), None, [f"Missing Q_ref.csv: {qref_path}"]

    with open(qref_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []
        cols_norm = [c.strip() for c in cols]
        alt_map = config.get("qref_alt_column_map", {})
        # map columns if needed
        def col_key(c: str) -> str:
            return alt_map.get(c, c)

        mapped = [col_key(c) for c in cols_norm]
        # build index
        idx = {m: i for i, m in enumerate(mapped)}
        if "t_s" not in idx or "Q_ml_s" not in idx:
            issues.append(f"Q_ref missing required columns. Found: {cols_norm} mapped: {mapped}")
            return np.array([]), np.array([]), None, issues

        t_list, q_list, v_list = [], [], []
        has_v = "V_ml" in idx
        for row in reader:
            t = safe_float(row.get(cols_norm[idx["t_s"]]) if "t_s" in idx else row.get("t_s"))
            q = safe_float(row.get(cols_norm[idx["Q_ml_s"]]) if "Q_ml_s" in idx else row.get("Q_ml_s"))
            if t is None or q is None:
                continue
            t_list.append(t)
            q_list.append(max(0.0, q))
            if has_v:
                v = safe_float(row.get(cols_norm[idx["V_ml"]]) if "V_ml" in idx else row.get("V_ml"))
                v_list.append(v if v is not None else math.nan)

        t = np.array(t_list, dtype=float)
        q = np.array(q_list, dtype=float)
        v = np.array(v_list, dtype=float) if has_v and len(v_list) == len(t_list) else None
        return t, q, v, issues


def check_qref_sanity(t: np.ndarray, q: np.ndarray) -> List[str]:
    issues = []
    if t.size < 3:
        issues.append("Q_ref too short (<3 samples)")
        return issues
    if np.any(np.diff(t) <= 0):
        issues.append("Q_ref time is not strictly increasing")
    if np.any(q < -1e-6):
        issues.append("Q_ref contains negative flow values")
    if np.nanmax(q) > 200:
        issues.append("Q_ref Qmax unusually high (>200 ml/s) - check units/export")
    return issues


def integrate_flow(t: np.ndarray, q: np.ndarray) -> float:
    if t.size < 2:
        return float("nan")
    trapezoid = getattr(np, "trapezoid", None)
    if trapezoid is not None:
        return float(trapezoid(q, t))
    return float(np.trapz(q, t))


def convert_audio_to_wav(audio_path: Path) -> Path:
    """
    Converts audio file to wav via ffmpeg into a temp file.
    If already wav returns the same path.
    """
    if audio_path.suffix.lower() == ".wav":
        return audio_path

    ffmpeg = shutil.which("ffmpeg")
    if ffmpeg is None:
        raise RuntimeError("ffmpeg not found in PATH; cannot convert audio.m4a -> wav")

    tmpdir = Path(tempfile.mkdtemp(prefix="uroflow_audio_"))
    out = tmpdir / "audio.wav"
    cmd = [ffmpeg, "-y", "-i", str(audio_path), "-ac", "1", "-ar", "48000", str(out)]
    subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=True)
    return out


def detect_audio_onset(wav_path: Path, config: dict) -> Tuple[Optional[float], dict, List[str]]:
    """
    Simple RMS-based onset detector.
    Returns onset_time_seconds from start of file.
    """
    issues = []
    debug = {}

    import wave
    try:
        with wave.open(str(wav_path), "rb") as wf:
            n_channels = wf.getnchannels()
            fs = wf.getframerate()
            n_frames = wf.getnframes()
            sampwidth = wf.getsampwidth()
            raw = wf.readframes(n_frames)
    except Exception as e:
        return None, {}, [f"Cannot read wav: {e}"]

    if n_channels != 1:
        issues.append(f"Audio has {n_channels} channels; expected mono")
    debug["fs"] = fs
    debug["n_frames"] = n_frames
    debug["sampwidth"] = sampwidth

    # decode int16/24/32 to float
    if sampwidth == 2:
        x = np.frombuffer(raw, dtype=np.int16).astype(np.float32) / 32768.0
    elif sampwidth == 3:
        # 24-bit little-endian
        a = np.frombuffer(raw, dtype=np.uint8).reshape(-1, 3)
        b = (a[:,0].astype(np.int32) | (a[:,1].astype(np.int32) << 8) | (a[:,2].astype(np.int32) << 16))
        # sign extend
        b = np.where(b & 0x800000, b | ~0xFFFFFF, b)
        x = b.astype(np.float32) / 8388608.0
    elif sampwidth == 4:
        x = np.frombuffer(raw, dtype=np.int32).astype(np.float32) / 2147483648.0
    else:
        issues.append(f"Unsupported sampwidth: {sampwidth}")
        return None, debug, issues

    win_ms = float(config.get("audio_onset_window_ms", 20))
    hop_ms = float(config.get("audio_onset_hop_ms", 10))
    win = max(1, int(fs * win_ms / 1000.0))
    hop = max(1, int(fs * hop_ms / 1000.0))

    # Compute RMS over frames
    rms = []
    times = []
    for start in range(0, len(x) - win, hop):
        seg = x[start:start+win]
        rms.append(float(np.sqrt(np.mean(seg*seg) + 1e-12)))
        times.append(start / fs)
    rms = np.array(rms)
    times = np.array(times)
    if rms.size < 5:
        return None, debug, ["Audio too short for onset detection"]

    # baseline from first 1 sec (or first 10% if shorter)
    baseline_mask = times <= min(1.0, times[-1] * 0.2)
    base = rms[baseline_mask]
    mu = float(np.mean(base))
    sigma = float(np.std(base) + 1e-6)
    z = (rms - mu) / sigma
    thr = float(config.get("audio_onset_zscore_threshold", 6.0))

    sustain_ms = float(config.get("audio_onset_min_sustain_ms", 200))
    sustain_frames = max(1, int(sustain_ms / hop_ms))
    onset_idx = None
    for i in range(len(z) - sustain_frames):
        if np.all(z[i:i+sustain_frames] >= thr):
            onset_idx = i
            break

    debug["baseline_mu"] = mu
    debug["baseline_sigma"] = sigma
    debug["z_thr"] = thr
    debug["sustain_frames"] = sustain_frames

    if onset_idx is None:
        issues.append("Audio onset not detected (below threshold)")
        return None, debug, issues

    onset_time = float(times[onset_idx])
    debug["onset_time_s"] = onset_time
    return onset_time, debug, issues


def audio_proxy_q(wav_path: Path, onset_time_s: float, target_hz: float = 10.0) -> Tuple[np.ndarray, np.ndarray, List[str]]:
    """
    Create a simple proxy flow signal from audio RMS energy after onset.
    Returns (t_s, proxy) with t_s starting at 0 (onset-aligned).
    """
    issues = []
    import wave
    with wave.open(str(wav_path), "rb") as wf:
        fs = wf.getframerate()
        n_frames = wf.getnframes()
        sampwidth = wf.getsampwidth()
        raw = wf.readframes(n_frames)

    if sampwidth != 2:
        issues.append("audio_proxy_q expects 16-bit wav; convert via ffmpeg -ac 1 -ar 48000")
        return np.array([]), np.array([]), issues

    x = np.frombuffer(raw, dtype=np.int16).astype(np.float32) / 32768.0
    start = int(onset_time_s * fs)
    if start >= len(x):
        return np.array([]), np.array([]), ["onset beyond audio length"]

    x = x[start:]
    # window of 100 ms
    win = int(fs * 0.1)
    hop = int(fs * 0.1)
    rms = []
    for i in range(0, len(x) - win, hop):
        seg = x[i:i+win]
        rms.append(float(np.sqrt(np.mean(seg*seg) + 1e-12)))
    if len(rms) < 3:
        return np.array([]), np.array([]), ["audio too short after onset"]

    proxy = np.array(rms, dtype=float)
    # normalize robustly
    p5, p95 = np.percentile(proxy, [5, 95])
    if p95 - p5 > 1e-9:
        proxy = (proxy - p5) / (p95 - p5)
    proxy = np.clip(proxy, 0, 2)
    t = np.arange(len(proxy)) / target_hz
    return t, proxy, issues


def resample_to_grid(t: np.ndarray, x: np.ndarray, grid_t: np.ndarray) -> np.ndarray:
    """
    Linear interpolation, assumes t increasing.
    """
    if t.size < 2 or x.size < 2:
        return np.full_like(grid_t, np.nan, dtype=float)
    return np.interp(grid_t, t, x, left=np.nan, right=np.nan)


def pearson_corr(a: np.ndarray, b: np.ndarray) -> Optional[float]:
    mask = np.isfinite(a) & np.isfinite(b)
    if mask.sum() < 5:
        return None
    aa = a[mask]
    bb = b[mask]
    if np.std(aa) < 1e-9 or np.std(bb) < 1e-9:
        return None
    return float(np.corrcoef(aa, bb)[0, 1])


def check_mp4_header(mp4_path: Path) -> Optional[str]:
    if not mp4_path.exists():
        return "missing"
    try:
        with open(mp4_path, "rb") as f:
            head = f.read(16)
        if b"ftyp" not in head:
            return "invalid_header"
    except Exception as e:
        return f"read_error:{e}"
    return None


def sha256_file(path: Path, block_size: int = 1 << 20) -> str:
    h = __import__("hashlib").sha256()
    with open(path, "rb") as f:
        while True:
            b = f.read(block_size)
            if not b:
                break
            h.update(b)
    return h.hexdigest()
