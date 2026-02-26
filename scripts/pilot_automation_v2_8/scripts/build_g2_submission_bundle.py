#!/usr/bin/env python3
"""
G2 Submission Bundle Builder (EU MDR + US FDA) – offline

What it does
- Reads the official submission folder indexes:
  * EU MDR Annex II/III folder index (xlsx)
  * US FDA submission folder index (xlsx)
- Copies all "Included"/"Required" items into a timestamped bundle folder
- Generates "EXECUTED" indexes with:
  Present (Y/N), SHA256, Bundle path
- Creates a missing-items report and an optional ZIP of the bundle.

This is meant to support Gate G2 (submission readiness) by making bundle assembly repeatable
and auditable.

Example
python scripts/build_g2_submission_bundle.py --submission_root ../../ --out_dir outputs/g2_bundle

"""

from __future__ import annotations

import argparse
import hashlib
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


DEFAULT_INCLUDE_STATUSES = {"Included", "Required"}


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk_size)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def read_index_rows(index_xlsx: Path) -> Tuple[List[str], List[dict]]:
    """
    Expects a sheet named 'Index' with header row:
      Section | File name | Relative path | Status | Notes
    Returns (headers, rows) where rows are dicts.
    """
    wb = load_workbook(index_xlsx)
    if "Index" not in wb.sheetnames:
        raise ValueError(f"No 'Index' sheet in {index_xlsx}")
    ws = wb["Index"]
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=c).value or "").strip())
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                empty = False
            row[h] = v
        if not empty:
            rows.append(row)
    return headers, rows


def write_executed_index(index_xlsx: Path, out_xlsx: Path, exec_rows: List[dict]) -> None:
    """
    Writes a copy of index_xlsx with extra columns:
      Present | SHA256 | Bundle path
    exec_rows must align to the original row order (excluding header).
    """
    wb = load_workbook(index_xlsx)
    ws = wb["Index"]
    # find new columns
    base_cols = ws.max_column
    col_present = base_cols + 1
    col_sha = base_cols + 2
    col_bundle = base_cols + 3

    ws.cell(row=1, column=col_present).value = "Present"
    ws.cell(row=1, column=col_sha).value = "SHA256"
    ws.cell(row=1, column=col_bundle).value = "Bundle path"

    # style header
    for c in range(1, col_bundle + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, er in enumerate(exec_rows, start=2):
        ws.cell(row=i, column=col_present).value = er.get("Present", "")
        ws.cell(row=i, column=col_sha).value = er.get("SHA256", "")
        ws.cell(row=i, column=col_bundle).value = er.get("BundlePath", "")

    ensure_parent(out_xlsx)
    wb.save(out_xlsx)


def build_bundle_for_index(
    submission_root: Path,
    index_xlsx: Path,
    bundle_root: Path,
    include_statuses: set,
) -> Tuple[Path, Path, dict]:
    """
    Returns:
      (executed_index_path, missing_report_path, summary_dict)
    """
    headers, rows = read_index_rows(index_xlsx)
    exec_rows = []
    missing = []
    included_count = 0
    present_count = 0

    for row in rows:
        rel = str(row.get("Relative path") or "").strip()
        status = str(row.get("Status") or "").strip()
        src = submission_root / rel if rel else None

        include = status in include_statuses
        present = bool(src and src.exists())
        sha = sha256_file(src) if present and include else ""
        bundle_path = ""

        if include:
            included_count += 1
            if present:
                present_count += 1
                dst = bundle_root / rel
                ensure_parent(dst)
                shutil.copy2(src, dst)
                bundle_path = str(dst.relative_to(bundle_root))

        if include and not present:
            missing.append(rel)

        exec_rows.append({
            "Present": "Y" if present else "N",
            "SHA256": sha,
            "BundlePath": bundle_path,
            "Status": status,
            "RelativePath": rel,
        })

    # write executed index
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    executed_index = bundle_root.parent / f"{index_xlsx.stem}_EXECUTED_{ts}.xlsx"
    write_executed_index(index_xlsx, executed_index, exec_rows)

    # missing report
    missing_report = bundle_root.parent / f"{index_xlsx.stem}_MISSING_{ts}.txt"
    with open(missing_report, "w", encoding="utf-8") as f:
        f.write(f"Missing items for bundle based on index: {index_xlsx.name}\n")
        f.write(f"Timestamp: {ts}\n\n")
        for m in missing:
            f.write(m + "\n")

    summary = {
        "index": str(index_xlsx),
        "included_items": included_count,
        "present_items": present_count,
        "missing_items": included_count - present_count,
        "missing_list_path": str(missing_report),
        "executed_index_path": str(executed_index),
    }
    return executed_index, missing_report, summary


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--submission_root", required=True, help="Path to Submission_Build_vX.Y folder")
    ap.add_argument("--out_dir", default="outputs/g2_bundle", help="Where to write bundles and reports")
    ap.add_argument("--eu_index", default="06_EU_MDR/Annex_II_III_Submission_Folder/EU_MDR_AnnexII_III_Submission_Folder_Index_v2.4.xlsx")
    ap.add_argument("--fda_index", default="07_US_FDA/FDA_Submission_Folder/FDA_Submission_Folder_Index_v2.2.xlsx")
    ap.add_argument("--include_statuses", default="Included,Required", help="Comma-separated statuses to include")
    ap.add_argument("--zip_bundle", action="store_true", help="Also ZIP the bundle folder")
    args = ap.parse_args()

    submission_root = Path(args.submission_root).resolve()
    out_dir = Path(args.out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    include_statuses = {s.strip() for s in args.include_statuses.split(",") if s.strip()}
    if not include_statuses:
        include_statuses = set(DEFAULT_INCLUDE_STATUSES)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bundle_root = out_dir / f"G2_Submission_Bundle_{ts}"
    bundle_root.mkdir(parents=True, exist_ok=True)

    reports_dir = out_dir / f"G2_Bundle_Reports_{ts}"
    reports_dir.mkdir(parents=True, exist_ok=True)

    summary_all = {"timestamp": ts, "submission_root": str(submission_root), "EU": None, "US": None}

    # EU
    eu_index = submission_root / args.eu_index
    if eu_index.exists():
        eu_bundle = bundle_root  # keep same root, preserve rel paths
        executed, missing, summary = build_bundle_for_index(submission_root, eu_index, eu_bundle, include_statuses)
        summary_all["EU"] = summary
    else:
        summary_all["EU"] = {"error": f"EU index not found: {eu_index}"}

    # US
    fda_index = submission_root / args.fda_index
    if fda_index.exists():
        fda_bundle = bundle_root
        executed, missing, summary = build_bundle_for_index(submission_root, fda_index, fda_bundle, include_statuses)
        summary_all["US"] = summary
    else:
        summary_all["US"] = {"error": f"FDA index not found: {fda_index}"}

    # Summary JSON
    summary_path = reports_dir / "g2_bundle_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        import json
        json.dump(summary_all, f, ensure_ascii=False, indent=2)

    # Optional zip
    if args.zip_bundle:
        zip_path = out_dir / f"G2_Submission_Bundle_{ts}.zip"
        shutil.make_archive(str(zip_path).replace(".zip",""), "zip", bundle_root)
        summary_all["bundle_zip"] = str(zip_path)

    print(f"OK: bundle={bundle_root}")
    print(f"Reports: {reports_dir}")
    print(f"Summary: {summary_path}")

if __name__ == "__main__":
    main()
