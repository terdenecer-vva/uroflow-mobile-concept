#!/usr/bin/env python3
"""
validate_artifacts_by_profile.py

Checks that each record in a golden dataset conforms to a selected data artifact profile (P0-P3).
- Ensures required artifacts exist
- Ensures forbidden raw artifacts are not present (privacy-by-default)

Usage:
  python validate_artifacts_by_profile.py --dataset_root <PATH> --manifest <manifest.csv|.xlsx> --profile P0
  python validate_artifacts_by_profile.py --dataset_root <PATH> --manifest <manifest.csv|.xlsx> --use_manifest_profile

Manifest expected columns:
  - record_id (required)
  - profile_id (optional, if --use_manifest_profile)
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Dict, List, Tuple

try:
    import openpyxl
except Exception:
    openpyxl = None


def load_config(cfg_path: Path) -> Dict:
    return json.loads(cfg_path.read_text(encoding="utf-8"))


def read_manifest(manifest_path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    if manifest_path.suffix.lower() == ".csv":
        with manifest_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append({k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in r.items()})
        return rows

    if manifest_path.suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        if openpyxl is None:
            raise RuntimeError("openpyxl is required to read xlsx manifests.")
        wb = openpyxl.load_workbook(manifest_path, data_only=True)
        # Prefer a sheet named 'manifest' if exists, else active
        ws = wb["manifest"] if "manifest" in wb.sheetnames else wb.active
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            r = {}
            for h, v in zip(headers, row):
                if not h:
                    continue
                r[h] = (str(v).strip() if v is not None else "")
            rows.append(r)
        return rows

    raise ValueError(f"Unsupported manifest format: {manifest_path}")


def glob_any(record_dir: Path, pattern: str) -> List[Path]:
    return list(record_dir.glob(pattern))


def validate_record(record_dir: Path, profile: Dict) -> Tuple[List[str], List[str], List[str]]:
    missing: List[str] = []
    present_forbidden: List[str] = []
    notes: List[str] = []

    # required patterns
    for pat in profile.get("required", []):
        hits = glob_any(record_dir, pat)
        if len(hits) == 0:
            missing.append(pat)

    # forbidden patterns
    for pat in profile.get("forbidden", []):
        hits = glob_any(record_dir, pat)
        if len(hits) > 0:
            # list a few examples
            present_forbidden.append(f"{pat} -> {hits[0].as_posix()}" + (" (+more)" if len(hits) > 1 else ""))

    # optional with warnings (recommendations)
    for pat in profile.get("recommended", []):
        hits = glob_any(record_dir, pat)
        if len(hits) == 0:
            notes.append(f"recommended_missing:{pat}")

    return missing, present_forbidden, notes


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset_root", required=True, type=Path)
    ap.add_argument("--manifest", required=True, type=Path)
    ap.add_argument("--config", default=Path(__file__).resolve().parent.parent / "config" / "data_artifact_profile_config.json", type=Path)
    ap.add_argument("--profile", choices=["P0", "P1", "P2", "P3"], help="Apply this profile to all records.")
    ap.add_argument("--use_manifest_profile", action="store_true", help="Use manifest column 'profile_id' per record.")
    ap.add_argument("--out_dir", default=Path("outputs/validate_artifacts"), type=Path)

    args = ap.parse_args()

    cfg = load_config(args.config)
    profiles = cfg.get("profiles", {})
    if args.use_manifest_profile and args.profile:
        raise ValueError("Choose either --profile or --use_manifest_profile, not both.")
    if not args.use_manifest_profile and not args.profile:
        raise ValueError("Provide --profile or --use_manifest_profile.")

    rows = read_manifest(args.manifest)
    args.out_dir.mkdir(parents=True, exist_ok=True)

    results = []
    any_missing = False
    any_forbidden = False

    for r in rows:
        record_id = r.get("record_id", "").strip()
        if not record_id:
            continue
        profile_id = args.profile
        if args.use_manifest_profile:
            profile_id = r.get("profile_id", "").strip() or cfg.get("default_profile", "P0")
        if profile_id not in profiles:
            raise ValueError(f"Unknown profile_id {profile_id} for record {record_id}")

        record_dir = args.dataset_root / "records" / record_id
        if not record_dir.exists():
            any_missing = True
            results.append({
                "record_id": record_id,
                "profile_id": profile_id,
                "status": "MISSING_RECORD_DIR",
                "missing": ["record_dir"],
                "forbidden_present": [],
                "notes": [],
            })
            continue

        missing, forbidden_present, notes = validate_record(record_dir, profiles[profile_id])
        status = "PASS"
        if missing:
            status = "FAIL_MISSING"
            any_missing = True
        if forbidden_present:
            status = "FAIL_FORBIDDEN" if status == "PASS" else status + "+FORBIDDEN"
            any_forbidden = True

        results.append({
            "record_id": record_id,
            "profile_id": profile_id,
            "status": status,
            "missing": missing,
            "forbidden_present": forbidden_present,
            "notes": notes,
        })

    # write outputs
    out_json = args.out_dir / "artifact_profile_validation.json"
    out_csv = args.out_dir / "artifact_profile_validation.csv"

    out_json.write_text(json.dumps({
        "dataset_root": args.dataset_root.as_posix(),
        "manifest": args.manifest.as_posix(),
        "config": args.config.as_posix(),
        "generated_at": cfg.get("generated_at", ""),
        "summary": {
            "total_records": len(results),
            "any_missing": any_missing,
            "any_forbidden": any_forbidden,
        },
        "results": results,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    with out_csv.open("w", encoding="utf-8", newline="") as f:
        fieldnames = ["record_id", "profile_id", "status", "missing", "forbidden_present", "notes"]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for it in results:
            w.writerow({
                "record_id": it["record_id"],
                "profile_id": it["profile_id"],
                "status": it["status"],
                "missing": ";".join(it["missing"]),
                "forbidden_present": ";".join(it["forbidden_present"]),
                "notes": ";".join(it["notes"]),
            })

    # exit code
    if any_missing or any_forbidden:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
