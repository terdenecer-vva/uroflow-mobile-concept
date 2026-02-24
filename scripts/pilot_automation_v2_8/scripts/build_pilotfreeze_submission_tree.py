#!/usr/bin/env python3
"""Build a *pilot-freeze* submission-ready tree (EU + US) without archives/duplicates.

The goal is a clean folder that can be handed to RA/QA for review:
- Includes only files referenced in EU Annex II/III index and US FDA index
- Optionally includes additional pilot-freeze records (claims lock, acceptance lock, freeze record, executed GSPR/master index)
- Generates SHA256 checksums for the resulting tree
- Produces a manifest + missing report

Usage (copy mode):
  python build_pilotfreeze_submission_tree.py --build_root <Submission_Build> \
    --out_dir <OUTPUT_DIR> \
    --eu_index 06_EU_MDR/Annex_II_III_Submission_Folder/EU_MDR_AnnexII_III_Submission_Folder_Index_v2.2.xlsx \
    --us_index 07_US_FDA/FDA_Submission_Folder/FDA_Submission_Folder_Index_v2.2.xlsx \
    --extra_list config/pilotfreeze_extra_includes.txt

Usage (dry run):
  python build_pilotfreeze_submission_tree.py ... --dry_run

Notes:
- The index files are expected to have columns: Section, File name, Relative path, Status
- Only rows with Status starting with 'Included' are included.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Dict, List, Set

import openpyxl

from uroflow_qa_utils import sha256_file


def _read_index(index_path: Path) -> List[str]:
    wb = openpyxl.load_workbook(index_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    header = [str(h).strip() if h is not None else "" for h in header]

    def col(name: str, fallback: int) -> int:
        return header.index(name) + 1 if name in header else fallback

    col_rel = col("Relative path", 3)
    col_status = col("Status", 4)

    rels = []
    for r in range(2, ws.max_row + 1):
        rel = ws.cell(r, col_rel).value
        status = ws.cell(r, col_status).value
        rel = str(rel).strip() if rel is not None else ""
        status = str(status).strip() if status is not None else ""
        if not rel:
            continue
        if not status.lower().startswith("included"):
            continue
        rels.append(rel)
    return rels


def _read_extra_list(extra_path: Path) -> List[str]:
    if not extra_path.exists():
        return []
    lines = []
    for line in extra_path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        lines.append(s)
    return lines


def _copy_file(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_bytes(src.read_bytes())


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--build_root", required=True)
    ap.add_argument("--out_dir", required=True)
    ap.add_argument("--eu_index", required=True)
    ap.add_argument("--us_index", required=True)
    ap.add_argument("--extra_list", default="")
    ap.add_argument("--dry_run", action="store_true")
    args = ap.parse_args()

    build_root = Path(args.build_root).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()

    eu_index = Path(args.eu_index)
    if not eu_index.is_absolute():
        eu_index = (build_root / eu_index).resolve()

    us_index = Path(args.us_index)
    if not us_index.is_absolute():
        us_index = (build_root / us_index).resolve()

    if not eu_index.exists():
        raise FileNotFoundError(f"EU index not found: {eu_index}")
    if not us_index.exists():
        raise FileNotFoundError(f"US index not found: {us_index}")

    required: List[str] = []
    required += _read_index(eu_index)
    required += _read_index(us_index)

    extra_list = []
    if args.extra_list:
        extra_path = Path(args.extra_list)
        if not extra_path.is_absolute():
            extra_path = (build_root / extra_path).resolve()
        extra_list = _read_extra_list(extra_path)
        required += extra_list

    # Deduplicate
    seen: Set[str] = set()
    ordered: List[str] = []
    for rel in required:
        rel = rel.strip().replace("\\", "/")
        if not rel:
            continue
        if rel in seen:
            continue
        # Avoid archive content by default
        if "/Archive/" in f"/{rel}/":
            continue
        seen.add(rel)
        ordered.append(rel)

    missing = []
    included = []

    if not args.dry_run:
        if out_dir.exists():
            # do not delete automatically; user may want to version outputs. We'll create anyway.
            pass
        out_dir.mkdir(parents=True, exist_ok=True)

    for rel in ordered:
        src = (build_root / rel).resolve()
        if src.exists() and src.is_file():
            included.append(rel)
            if not args.dry_run:
                dst = (out_dir / rel)
                _copy_file(src, dst)
        else:
            missing.append(rel)

    # Write manifest and missing report
    manifest_dir = out_dir if not args.dry_run else out_dir
    manifest_dir.mkdir(parents=True, exist_ok=True)

    (manifest_dir / "pilotfreeze_tree_manifest.txt").write_text(
        "\n".join(included) + "\n", encoding="utf-8"
    )

    with open(manifest_dir / "pilotfreeze_tree_missing.csv", "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["rel_path", "exists"])
        for rel in missing:
            w.writerow([rel, "N"])

    # Checksums (only if copied)
    if not args.dry_run:
        lines = []
        for p in sorted(out_dir.rglob("*")):
            if p.is_file():
                h = sha256_file(p)
                relp = p.relative_to(out_dir).as_posix()
                lines.append(f"{h}  {relp}")
        (out_dir / "checksums.sha256").write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"[OK] Included: {len(included)} | Missing: {len(missing)} | Dry-run: {args.dry_run}")
    print(f"[OK] Manifest: {manifest_dir / 'pilotfreeze_tree_manifest.txt'}")
    print(f"[OK] Missing report: {manifest_dir / 'pilotfreeze_tree_missing.csv'}")
    if not args.dry_run:
        print(f"[OK] Checksums: {out_dir / 'checksums.sha256'}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
