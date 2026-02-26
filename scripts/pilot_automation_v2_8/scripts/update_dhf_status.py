#!/usr/bin/env python3
"""Update DHF Index status workbook with *executed* file-level checks.

Goal
- Turn the DHF Index into an auditable register by adding:
  - Exists (Y/N)
  - Resolved path (relative)
  - SHA256 short (basename:hash12)
  - Last checked timestamp

This script does NOT change the human-controlled columns (Status/Approved date/etc.).

Usage
  python update_dhf_status.py \
    --build_root <Submission_Build_vX.X> \
    --dhf_in 01_Product_QMS/Uroflow_DHF_Index_Status_Register_v1.2.xlsx \
    --dhf_out 01_Product_QMS/Uroflow_DHF_Index_Status_Register_v1.3_EXECUTED_AUTO.xlsx
"""

from __future__ import annotations

import argparse
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

from uroflow_qa_utils import sha256_file


SKIP_DIR_NAMES = {
    "Archive",
    ".git",
    "__pycache__",
    "records",
    "raw",
    "videos",
    "audio",
}

SEARCH_ROOTS_DEFAULT = [
    "01_Product_QMS",
    "02_Algorithm_and_ML",
    "04_Bench_Testing",
    "05_Clinical",
    "06_EU_MDR",
    "07_US_FDA",
    "08_Privacy_and_Security",
    "09_HFE_Usability",
    "10_Pilot_Automation",
]


def _ensure_columns(ws, headers: List[str]) -> List[int]:
    existing = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    existing = [str(v).strip() if v is not None else "" for v in existing]

    cols = []
    next_col = ws.max_column + 1
    for h in headers:
        if h in existing:
            cols.append(existing.index(h) + 1)
        else:
            ws.cell(1, next_col).value = h
            cols.append(next_col)
            next_col += 1
    return cols


def _walk_find_by_name(root: Path, filename: str) -> Optional[Path]:
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in SKIP_DIR_NAMES]
        if filename in filenames:
            return Path(dirpath) / filename
    return None


def _find_file(build_root: Path, filename: str, search_roots: List[Path]) -> Optional[Path]:
    for sr in search_roots:
        if not sr.exists() or not sr.is_dir():
            continue
        p = _walk_find_by_name(sr, filename)
        if p is not None:
            return p
    # fallback: whole build (still pruned)
    return _walk_find_by_name(build_root, filename)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--build_root", required=True)
    ap.add_argument("--dhf_in", required=True)
    ap.add_argument("--dhf_out", required=True)
    ap.add_argument(
        "--search_roots",
        default=";".join(SEARCH_ROOTS_DEFAULT),
        help="Semicolon-separated search roots (relative to build_root)",
    )
    args = ap.parse_args()

    build_root = Path(args.build_root).expanduser().resolve()

    dhf_in = Path(args.dhf_in)
    if not dhf_in.is_absolute():
        dhf_in = (build_root / dhf_in).resolve()

    dhf_out = Path(args.dhf_out)
    if not dhf_out.is_absolute():
        dhf_out = (build_root / dhf_out).resolve()

    if not dhf_in.exists():
        raise FileNotFoundError(f"DHF input not found: {dhf_in}")

    # Build search roots
    search_roots: List[Path] = []
    for part in str(args.search_roots).split(";"):
        part = part.strip()
        if not part:
            continue
        p = (build_root / part).resolve()
        if p.exists() and p.is_dir():
            search_roots.append(p)

    wb = openpyxl.load_workbook(dhf_in)
    if "DHF_Index" not in wb.sheetnames:
        raise ValueError("Sheet 'DHF_Index' not found")

    ws = wb["DHF_Index"]

    # Identify input columns
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header = [str(h).strip() if h is not None else "" for h in header]

    def col(name: str) -> int:
        try:
            return header.index(name) + 1
        except ValueError:
            return -1

    c_file = col("File name")
    c_path = col("Build path")

    if c_file < 0:
        raise ValueError("'File name' column not found")

    # Add autofill columns
    new_cols = _ensure_columns(
        ws,
        [
            "Autofill: Exists (Y/N)",
            "Autofill: Resolved path (relative)",
            "Autofill: SHA256 short (basename:hash12)",
            "Autofill: Last checked (UTC)",
        ],
    )
    c_exists, c_resolved, c_sha, c_ts = new_cols

    total = 0
    present = 0
    missing = 0

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")

    for r in range(2, ws.max_row + 1):
        fn = ws.cell(r, c_file).value
        if fn is None or str(fn).strip() == "":
            continue
        fn = str(fn).strip()

        build_path = ws.cell(r, c_path).value if c_path > 0 else None
        build_path = str(build_path).strip() if build_path is not None else ""

        resolved: Optional[Path] = None

        # 1) Use build path if provided
        if build_path:
            p = Path(build_path)
            if not p.is_absolute():
                p = (build_root / p).resolve()
            resolved = p
            if not resolved.exists():
                # fallback: search by name
                resolved = _find_file(build_root, fn, search_roots)
        else:
            resolved = _find_file(build_root, fn, search_roots)

        total += 1

        if resolved is not None and resolved.exists() and resolved.is_file():
            present += 1
            ws.cell(r, c_exists).value = "Y"
            try:
                rel = str(resolved.relative_to(build_root))
            except Exception:
                rel = str(resolved)
            ws.cell(r, c_resolved).value = rel
            try:
                h = sha256_file(resolved)
                ws.cell(r, c_sha).value = f"{Path(rel).name}:{h[:12]}"
            except Exception:
                ws.cell(r, c_sha).value = f"{Path(rel).name}:sha_error"
        else:
            missing += 1
            ws.cell(r, c_exists).value = "N"
            ws.cell(r, c_resolved).value = ""
            ws.cell(r, c_sha).value = ""

        ws.cell(r, c_ts).value = ts

    # Summary sheet
    if "Autofill_Summary" in wb.sheetnames:
        del wb["Autofill_Summary"]
    ws_sum = wb.create_sheet("Autofill_Summary")
    ws_sum["A1"].value = "Build root"
    ws_sum["B1"].value = str(build_root)
    ws_sum["A2"].value = "DHF input"
    ws_sum["B2"].value = str(dhf_in)
    ws_sum["A3"].value = "Checked at (UTC)"
    ws_sum["B3"].value = ts
    ws_sum["A5"].value = "Total entries checked"
    ws_sum["B5"].value = total
    ws_sum["A6"].value = "Present"
    ws_sum["B6"].value = present
    ws_sum["A7"].value = "Missing"
    ws_sum["B7"].value = missing

    dhf_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dhf_out)
    print(f"[OK] Wrote: {dhf_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
