#!/usr/bin/env python3
"""
Golden dataset → CSR-style TFL generator (offline)

What it does
- Reads the golden dataset (record folders) + manifest (CSV/XLSX)
- Extracts reference metrics (from manifest and/or Q_ref.csv)
- Extracts app/predicted metrics (from Q_pred.csv and/or app_result.json)
- Generates:
  1) Record-level listing (CSV)
  2) Summary stats (JSON)
  3) Bland–Altman plots (PNG) [optional]
  4) Filled CSR TFL Excel template (DATA_IMPORT sheet) [optional]
  5) Short PDF summary report (optional)

Inputs
- dataset_root: folder that contains records/{record_id}/
- manifest: CSV or XLSX (same as QA scripts)

Notes
- Designed to work offline.
- Uses existing QA utilities for manifest loading and Q_ref parsing.
"""

from __future__ import annotations

import argparse
import json
import math
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

from uroflow_qa_utils import load_manifest, find_record_folder, parse_qref_csv, integrate_flow, safe_float


def load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(obj: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)


def _norm_str(x: str) -> str:
    return (x or "").strip()


def _map_bool(x: str, mapping: Dict[str, str], default: str = "") -> str:
    s = _norm_str(str(x)).lower()
    if s in mapping:
        return mapping[s]
    s2 = _norm_str(str(x))
    if s2 in mapping:
        return mapping[s2]
    return default


def _map_artifact(x: str, mapping: Dict[str, int], default: int = 0) -> int:
    s = _norm_str(str(x)).lower()
    if s in mapping:
        return int(mapping[s])
    s2 = _norm_str(str(x))
    if s2 in mapping:
        return int(mapping[s2])
    return int(default)


def _read_curve_csv(path: Path, alt_map: Dict[str, str]) -> Tuple[np.ndarray, np.ndarray]:
    """
    Expects columns: t_s, Q_ml_s (or alternatives mapped by alt_map)
    """
    if not path.exists():
        return np.array([]), np.array([])
    import csv

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return np.array([]), np.array([])
        # normalize columns
        cols = [c.strip() for c in reader.fieldnames]
        mapped = [alt_map.get(c, c) for c in cols]
        idx = {mapped[i]: cols[i] for i in range(len(cols))}
        if "t_s" not in idx or "Q_ml_s" not in idx:
            return np.array([]), np.array([])

        t_list, q_list = [], []
        for row in reader:
            t = safe_float(row.get(idx["t_s"]))
            q = safe_float(row.get(idx["Q_ml_s"]))
            if t is None or q is None:
                continue
            t_list.append(float(t))
            q_list.append(max(0.0, float(q)))
        return np.array(t_list, dtype=float), np.array(q_list, dtype=float)


def compute_metrics_from_curve(t: np.ndarray, q: np.ndarray, flow_threshold: float = 0.5) -> Dict[str, float]:
    """
    Computes basic uroflow metrics from a flow curve.
    - t_start: first t where q>=threshold
    - t_end: last t where q>=threshold
    - flow_time: t_end - t_start
    - Qmax: max q in [t_start, t_end]
    - Vvoid: integral q dt in [t_start, t_end]
    - Qavg: Vvoid / flow_time
    """
    out = {
        "t_start_s": math.nan,
        "t_end_s": math.nan,
        "flow_time_s": math.nan,
        "Qmax_ml_s": math.nan,
        "Vvoid_ml": math.nan,
        "Qavg_ml_s": math.nan,
    }
    if t.size < 3 or q.size != t.size:
        return out

    # ensure sorted by t
    order = np.argsort(t)
    t = t[order]
    q = q[order]

    mask = q >= flow_threshold
    if not np.any(mask):
        # fallback: use any positive flow
        mask = q > 0.0
    if not np.any(mask):
        return out

    i0 = int(np.argmax(mask))
    i1 = int(len(mask) - 1 - np.argmax(mask[::-1]))

    t0 = float(t[i0])
    t1 = float(t[i1])
    if t1 <= t0:
        return out

    t_seg = t[i0:i1 + 1]
    q_seg = q[i0:i1 + 1]
    trapezoid = getattr(np, "trapezoid", None)
    if trapezoid is not None:
        v = float(trapezoid(q_seg, t_seg))
    else:
        v = float(np.trapz(q_seg, t_seg))
    qmax = float(np.max(q_seg))
    flow_time = float(t1 - t0)
    qavg = float(v / flow_time) if flow_time > 0 else math.nan

    out.update({
        "t_start_s": t0,
        "t_end_s": t1,
        "flow_time_s": flow_time,
        "Qmax_ml_s": qmax,
        "Vvoid_ml": v,
        "Qavg_ml_s": qavg,
    })
    return out


def ba_stats(ref: np.ndarray, pred: np.ndarray) -> Dict[str, float]:
    """
    Bland–Altman stats + MAE/MAPE
    """
    m = np.isfinite(ref) & np.isfinite(pred)
    ref = ref[m]
    pred = pred[m]
    if ref.size < 3:
        return {
            "n": int(ref.size),
            "bias": math.nan,
            "sd": math.nan,
            "loa_low": math.nan,
            "loa_high": math.nan,
            "mae": math.nan,
            "mape": math.nan,
        }
    diff = pred - ref
    bias = float(np.mean(diff))
    sd = float(np.std(diff, ddof=1))
    loa_low = bias - 1.96 * sd
    loa_high = bias + 1.96 * sd
    mae = float(np.mean(np.abs(diff)))
    # avoid divide by zero
    denom = np.where(np.abs(ref) < 1e-9, np.nan, ref)
    mape = float(np.nanmean(np.abs(diff / denom)) * 100.0)
    return {
        "n": int(ref.size),
        "bias": bias,
        "sd": sd,
        "loa_low": float(loa_low),
        "loa_high": float(loa_high),
        "mae": mae,
        "mape": mape,
    }


def try_make_ba_plot(x_mean: np.ndarray, y_diff: np.ndarray, stats: Dict[str, float], title: str, out_png: Path) -> None:
    try:
        import matplotlib.pyplot as plt
    except Exception:
        return

    out_png.parent.mkdir(parents=True, exist_ok=True)

    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.scatter(x_mean, y_diff, s=12)
    if math.isfinite(stats.get("bias", math.nan)):
        ax.axhline(stats["bias"])
    if math.isfinite(stats.get("loa_low", math.nan)):
        ax.axhline(stats["loa_low"], linestyle="--")
    if math.isfinite(stats.get("loa_high", math.nan)):
        ax.axhline(stats["loa_high"], linestyle="--")
    ax.set_title(title)
    ax.set_xlabel("Mean of (ref, app)")
    ax.set_ylabel("Difference (app - ref)")
    fig.tight_layout()
    fig.savefig(out_png, dpi=200)
    plt.close(fig)


def fill_csr_template(template_path: Path, out_path: Path, rows: List[dict]) -> None:
    """
    Fills DATA_IMPORT sheet in CSR TFL workbook template.
    Excel formulas will recalculate when opened in Excel.
    """
    import openpyxl

    wb = openpyxl.load_workbook(template_path)
    ws = wb["DATA_IMPORT"]
    header = [c.value for c in ws[1]]
    # clear existing rows (from 2 to max)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(header) + 1):
            ws.cell(r, c).value = None

    # write new rows
    for i, row in enumerate(rows, start=2):
        for j, key in enumerate(header, start=1):
            ws.cell(i, j).value = row.get(key, None)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset_root", required=True, help="Path to dataset root (contains records/)")
    ap.add_argument("--manifest", required=True, help="Path to manifest CSV/XLSX")
    ap.add_argument("--config", default=str(Path(__file__).resolve().parents[1] / "config" / "metrics_config.json"),
                    help="metrics_config.json")
    ap.add_argument("--out_dir", default=str(Path(__file__).resolve().parents[1] / "outputs" / "tfl"),
                    help="Output directory")
    ap.add_argument("--csr_template", default="", help="Path to CSR TFL workbook template (xlsx). If empty, will try 05_Clinical template.")
    ap.add_argument("--csr_out", default="", help="Output path for filled CSR workbook. If empty, out_dir will be used.")
    ap.add_argument("--make_plots", action="store_true", help="Generate BA plots (PNG)")
    ap.add_argument("--make_pdf", action="store_true", help="Generate a short PDF summary report (requires reportlab)")
    args = ap.parse_args()

    dataset_root = Path(args.dataset_root).expanduser().resolve()
    manifest_path = Path(args.manifest).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    cfg = load_json(Path(args.config))
    alt_map = cfg.get("curve_alt_column_map", {})
    flow_thr = float(cfg.get("flow_threshold_ml_s", 0.5))
    posture_map = cfg.get("posture_map", {})
    bool_map = cfg.get("bool_map", {})
    art_map = cfg.get("artifact_map", {})
    valid_map = cfg.get("valid_map", {})

    rows = load_manifest(manifest_path)

    listing = []
    csr_rows = []

    for r in rows:
        record_id = _norm_str(r.get("record_id", ""))
        if not record_id:
            continue
        rec_folder = find_record_folder(dataset_root, record_id, {"record_folder_candidates": ["records/{record_id}", "{record_id}"]})
        if rec_folder is None:
            # still include row with missing flags
            rec_folder = dataset_root / "records" / record_id

        # Reference: prefer manifest values if present, else compute from Q_ref.csv
        ref_qmax = safe_float(r.get("Qmax_ref_ml_s"))
        ref_qavg = safe_float(r.get("Qavg_ref_ml_s"))
        ref_v = safe_float(r.get("Vvoid_ref_ml"))
        ref_flow_time = safe_float(r.get("flow_time_ref_s"))

        t_ref, q_ref, _, _ = parse_qref_csv(rec_folder / "Q_ref.csv", {"qref_alt_column_map": alt_map})
        if (ref_qmax is None or ref_v is None or ref_flow_time is None) and t_ref.size >= 3:
            ref_m = compute_metrics_from_curve(t_ref, q_ref, flow_threshold=flow_thr)
            ref_qmax = ref_qmax if ref_qmax is not None else ref_m["Qmax_ml_s"]
            ref_v = ref_v if ref_v is not None else ref_m["Vvoid_ml"]
            ref_flow_time = ref_flow_time if ref_flow_time is not None else ref_m["flow_time_s"]
            ref_qavg = ref_qavg if ref_qavg is not None else ref_m["Qavg_ml_s"]

        # Predicted/app: Q_pred.csv → metrics; fallback to app_result.json; fallback to manifest app fields if exist
        pred_qmax = safe_float(r.get("Qmax_app_ml_s")) or safe_float(r.get("Qmax_pred_ml_s"))
        pred_qavg = safe_float(r.get("Qavg_app_ml_s")) or safe_float(r.get("Qavg_pred_ml_s"))
        pred_v = safe_float(r.get("Vvoid_app_ml")) or safe_float(r.get("Vvoid_pred_ml"))
        pred_flow_time = safe_float(r.get("flow_time_app_s")) or safe_float(r.get("flow_time_pred_s"))

        t_pred, q_pred = np.array([]), np.array([])
        # candidates
        for name in cfg.get("pred_curve_candidates", ["Q_pred.csv"]):
            p = rec_folder / name
            if p.exists():
                t_pred, q_pred = _read_curve_csv(p, alt_map)
                break

        if t_pred.size >= 3:
            pm = compute_metrics_from_curve(t_pred, q_pred, flow_threshold=flow_thr)
            pred_qmax = pred_qmax if pred_qmax is not None else pm["Qmax_ml_s"]
            pred_v = pred_v if pred_v is not None else pm["Vvoid_ml"]
            pred_flow_time = pred_flow_time if pred_flow_time is not None else pm["flow_time_s"]
            pred_qavg = pred_qavg if pred_qavg is not None else pm["Qavg_ml_s"]
        else:
            for name in cfg.get("pred_result_candidates", ["app_result.json"]):
                p = rec_folder / name
                if p.exists():
                    d = load_json(p)
                    pred_qmax = pred_qmax if pred_qmax is not None else safe_float(d.get("Qmax_ml_s"))
                    pred_qavg = pred_qavg if pred_qavg is not None else safe_float(d.get("Qavg_ml_s"))
                    pred_v = pred_v if pred_v is not None else safe_float(d.get("Vvoid_ml"))
                    pred_flow_time = pred_flow_time if pred_flow_time is not None else safe_float(d.get("flow_time_s"))
                    break

        # quality score
        qscore = safe_float(r.get("quality_score"))
        qjson = rec_folder / "quality.json"
        if qscore is None and qjson.exists():
            try:
                qd = load_json(qjson)
                qscore = safe_float(qd.get("quality_score"))
            except Exception:
                pass
        # allow meta.json
        if qscore is None:
            meta = rec_folder / "meta.json"
            if meta.exists():
                try:
                    md = load_json(meta)
                    qscore = safe_float(md.get("quality_score"))
                except Exception:
                    pass

        # Valid flag for primary
        valid = _map_bool(r.get("overall_record_valid", ""), valid_map, default="Y")
        # artifacts
        art_flush = _map_artifact(r.get("art_flush", "no"), art_map, default=0)
        art_motion = _map_artifact(r.get("art_phone_motion", "no"), art_map, default=0)
        art_not_in_water = _map_artifact(r.get("art_not_in_water", "no"), art_map, default=0)

        posture_raw = _norm_str(r.get("posture",""))
        posture = posture_map.get(posture_raw, posture_raw)

        # CSR template row
        csr_row = {
            "record_id": record_id,
            "subject_id": _norm_str(r.get("subject_id","")),
            "site_id": _norm_str(r.get("site_id","")),
            "toilet_id": _norm_str(r.get("toilet_id","")),
            "sex": _norm_str(r.get("sex","")),
            "posture": posture,
            "age": safe_float(r.get("age_years")),
            "noise_level": _norm_str(r.get("noise_level","")).upper(),
            "lighting": _norm_str(r.get("lighting","")).upper(),
            "lidar_available": _map_bool(r.get("lidar_available",""), bool_map, default=""),
            "toilet_scan": _map_bool(r.get("toilet_scan",""), bool_map, default=""),
            "pour_calibration": _map_bool(r.get("pour_calibration",""), bool_map, default=""),
            "quality_score": qscore,
            "valid_for_primary": valid,
            "artifact_flush": art_flush,
            "artifact_motion": art_motion,
            "artifact_not_in_water": art_not_in_water,
            "ref_Qmax_ml_s": ref_qmax,
            "app_Qmax_ml_s": pred_qmax,
            "ref_Qavg_ml_s": ref_qavg,
            "app_Qavg_ml_s": pred_qavg,
            "ref_Vvoid_ml": ref_v,
            "app_Vvoid_ml": pred_v,
            "ref_FlowTime_s": ref_flow_time,
            "app_FlowTime_s": pred_flow_time,
        }
        csr_rows.append(csr_row)

        # listing row (more verbose)
        listing.append({
            "record_id": record_id,
            "site_id": csr_row["site_id"],
            "toilet_id": csr_row["toilet_id"],
            "sex": csr_row["sex"],
            "posture": csr_row["posture"],
            "noise_level": csr_row["noise_level"],
            "lidar_available": csr_row["lidar_available"],
            "quality_score": qscore,
            "valid_for_primary": valid,
            "ref_Qmax_ml_s": ref_qmax,
            "app_Qmax_ml_s": pred_qmax,
            "ref_Qavg_ml_s": ref_qavg,
            "app_Qavg_ml_s": pred_qavg,
            "ref_Vvoid_ml": ref_v,
            "app_Vvoid_ml": pred_v,
            "ref_FlowTime_s": ref_flow_time,
            "app_FlowTime_s": pred_flow_time,
        })

    # Save listing CSV
    listing_csv = out_dir / "tfl_record_level.csv"
    if pd is not None:
        pd.DataFrame(listing).to_csv(listing_csv, index=False)
    else:
        import csv
        with open(listing_csv, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(listing[0].keys()) if listing else [])
            w.writeheader()
            for row in listing:
                w.writerow(row)

    # Compute overall BA stats for valid rows
    if pd is not None and listing:
        df = pd.DataFrame(listing)
        dfv = df[df["valid_for_primary"] == "Y"].copy()

        def _col(name: str) -> np.ndarray:
            return dfv[name].astype(float).to_numpy()

        metrics = {
            "Qmax": ("ref_Qmax_ml_s", "app_Qmax_ml_s"),
            "Qavg": ("ref_Qavg_ml_s", "app_Qavg_ml_s"),
            "Vvoid": ("ref_Vvoid_ml", "app_Vvoid_ml"),
            "FlowTime": ("ref_FlowTime_s", "app_FlowTime_s"),
        }
        summary = {"version": "1.1", "n_total": int(df.shape[0]), "n_valid": int(dfv.shape[0]), "metrics": {}}

        plots_dir = out_dir / "ba_plots"
        plots_dir.mkdir(parents=True, exist_ok=True)

        for mname, (rcol, pcol) in metrics.items():
            ref = dfv[rcol].astype(float).to_numpy()
            pred = dfv[pcol].astype(float).to_numpy()
            st = ba_stats(ref, pred)
            summary["metrics"][mname] = st

            if args.make_plots:
                xmean = (ref + pred) / 2.0
                ydiff = pred - ref
                try_make_ba_plot(xmean, ydiff, st, f"Bland–Altman: {mname}", plots_dir / f"BA_{mname}.png")

        save_json(summary, out_dir / "tfl_summary.json")

        # Optional: PDF summary
        if args.make_pdf:
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.units import mm
                from reportlab.pdfgen import canvas
                from reportlab.lib.utils import ImageReader
            except Exception:
                pass
            else:
                pdf_path = out_dir / "tfl_summary_report.pdf"
                c = canvas.Canvas(str(pdf_path), pagesize=A4)
                w, h = A4
                y = h - 20*mm
                c.setFont("Helvetica-Bold", 12)
                c.drawString(20*mm, y, "Golden dataset TFL summary (auto-generated)")
                y -= 8*mm
                c.setFont("Helvetica", 9)
                c.drawString(20*mm, y, f"Records: total={summary['n_total']} valid={summary['n_valid']}")
                y -= 10*mm

                # table-like text
                for mname, st in summary["metrics"].items():
                    line = f"{mname}: n={st['n']} bias={st['bias']:.3f} sd={st['sd']:.3f} LoA=[{st['loa_low']:.3f}; {st['loa_high']:.3f}] MAE={st['mae']:.3f} MAPE={st['mape']:.2f}%"
                    c.drawString(20*mm, y, line)
                    y -= 6*mm
                    if y < 30*mm:
                        c.showPage()
                        y = h - 20*mm

                # embed plots if present
                if args.make_plots:
                    for mname in metrics.keys():
                        imgp = plots_dir / f"BA_{mname}.png"
                        if imgp.exists():
                            if y < 80*mm:
                                c.showPage()
                                y = h - 20*mm
                            c.setFont("Helvetica-Bold", 10)
                            c.drawString(20*mm, y, f"{mname} BA plot")
                            y -= 5*mm
                            try:
                                img = ImageReader(str(imgp))
                                c.drawImage(img, 20*mm, y-60*mm, width=160*mm, height=60*mm, preserveAspectRatio=True, mask='auto')
                            except Exception:
                                pass
                            y -= 70*mm

                c.save()

    # Fill CSR template if requested
    csr_template = args.csr_template
    if not csr_template:
        # try to locate template relative to dataset_root (Submission build)
        # dataset_root likely points to a dataset, not submission build; so we don't guess too hard.
        csr_template = ""

    if csr_template:
        tpl = Path(csr_template).expanduser().resolve()
        if tpl.exists():
            out_xlsx = Path(args.csr_out).expanduser().resolve() if args.csr_out else (out_dir / "Uroflow_CSR_TFL_Workbook_filled.xlsx")
            fill_csr_template(tpl, out_xlsx, csr_rows)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
