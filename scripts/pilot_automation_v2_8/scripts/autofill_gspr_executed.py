#!/usr/bin/env python3
"""Autofill EU MDR GSPR checklist with file existence, SHA256, and Evidence_ID matches.

This script is intended to make the GSPR checklist *executed* and auditable:
- Verifies that each referenced file exists in the current submission build
- Computes SHA256 (short) for traceability
- Attempts to map referenced files to Evidence_IDs from the Pilot Readiness Evidence Manifest

Usage:
  python autofill_gspr_executed.py --build_root <Submission_Build> \
    --gspr_in 06_EU_MDR/Uroflow_EU_MDR_GSPR_Checklist_AnnexI_v1.1_EXECUTED.xlsx \
    --gspr_out 06_EU_MDR/Uroflow_EU_MDR_GSPR_Checklist_AnnexI_v1.2_EXECUTED_AUTO.xlsx \
    --evidence_manifest 05_Clinical/Uroflow_P0_Pilot_Readiness_Evidence_Manifest_v1.0.xlsx

Notes:
- Keeps original columns intact.
- Adds extra columns at the end of the GSPR_Checklist sheet.
"""

from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

from uroflow_qa_utils import sha256_file


def _parse_refs(cell_value) -> List[str]:
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    # split by newlines
    refs = [r.strip() for r in s.splitlines()]
    return [r for r in refs if r]


def _load_evidence_map(evidence_manifest_path: Path) -> Dict[str, str]:
    """Map expected file name -> Evidence_ID."""
    if not evidence_manifest_path.exists():
        return {}

    wb = openpyxl.load_workbook(evidence_manifest_path, data_only=True)
    sheet = "Manifest" if "Manifest" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    header = [c.value for c in ws[1]]
    header = [str(h).strip() if h is not None else "" for h in header]

    def idx(name: str) -> int:
        try:
            return header.index(name) + 1
        except ValueError:
            return -1

    col_eid = idx("Evidence_ID")
    col_expected = idx("Expected file name")
    if col_eid < 0 or col_expected < 0:
        return {}

    mapping: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        eid = ws.cell(r, col_eid).value
        exp = ws.cell(r, col_expected).value
        if eid is None or exp is None:
            continue
        eid = str(eid).strip()
        exp = str(exp).strip()
        if not eid or not exp:
            continue
        mapping[exp] = eid
    return mapping


def _ensure_columns(ws, headers: List[str]) -> Tuple[int, List[int]]:
    """Ensure headers exist at the end of row 1. Returns start_col and list of col indices."""
    existing = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    existing = [str(v).strip() if v is not None else "" for v in existing]

    start_col = ws.max_column + 1
    cols = []

    for h in headers:
        if h in existing:
            cols.append(existing.index(h) + 1)
        else:
            ws.cell(1, start_col).value = h
            cols.append(start_col)
            start_col += 1

    return (ws.max_column - len([h for h in headers if h not in existing]) + 1, cols)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--build_root", required=True, help="Path to Submission_Build root")
    ap.add_argument("--gspr_in", required=True, help="Input GSPR XLSX (relative to build_root or absolute)")
    ap.add_argument("--gspr_out", required=True, help="Output GSPR XLSX")
    ap.add_argument("--evidence_manifest", default="", help="Evidence manifest XLSX")
    args = ap.parse_args()

    build_root = Path(args.build_root).expanduser().resolve()
    gspr_in = Path(args.gspr_in)
    if not gspr_in.is_absolute():
        gspr_in = (build_root / gspr_in).resolve()

    gspr_out = Path(args.gspr_out)
    if not gspr_out.is_absolute():
        gspr_out = (build_root / gspr_out).resolve()

    evidence_manifest = Path(args.evidence_manifest) if args.evidence_manifest else Path()
    if evidence_manifest and not evidence_manifest.is_absolute():
        evidence_manifest = (build_root / evidence_manifest).resolve()

    if not gspr_in.exists():
        raise FileNotFoundError(f"GSPR input not found: {gspr_in}")

    evidence_map = _load_evidence_map(evidence_manifest) if evidence_manifest else {}

    wb = openpyxl.load_workbook(gspr_in)
    if "GSPR_Checklist" not in wb.sheetnames:
        raise ValueError("Sheet 'GSPR_Checklist' not found")

    ws = wb["GSPR_Checklist"]

    # Column I = 9 is 'Build file references (relative paths)' in current template
    COL_REFS = 9

    # Add autofill columns
    headers = [
        "Autofill: Present/Total refs",
        "Autofill: Missing refs (relative paths)",
        "Autofill: SHA256 short (basename:hash12)",
        "Autofill: Evidence_ID matches",
        "Autofill: Completeness (ALL_PRESENT/PARTIAL/ALL_MISSING/NO_REFS)",
    ]
    _, cols = _ensure_columns(ws, headers)
    col_present_total, col_missing, col_sha, col_evid, col_comp = cols

    missing_counter = Counter()
    total_refs = 0
    total_present = 0

    for r in range(2, ws.max_row + 1):
        refs = _parse_refs(ws.cell(r, COL_REFS).value)
        if not refs:
            ws.cell(r, col_present_total).value = ""
            ws.cell(r, col_missing).value = ""
            ws.cell(r, col_sha).value = ""
            ws.cell(r, col_evid).value = ""
            ws.cell(r, col_comp).value = "NO_REFS"
            continue

        total_refs += len(refs)
        present = 0
        missing: List[str] = []
        sha_entries: List[str] = []
        ev_ids = set()

        for rel in refs:
            p = (build_root / rel).resolve()
            if p.exists() and p.is_file():
                present += 1
                total_present += 1
                h = sha256_file(p)
                sha_entries.append(f"{Path(rel).name}:{h[:12]}")
                eid = evidence_map.get(Path(rel).name)
                if eid:
                    ev_ids.add(eid)
            else:
                missing.append(rel)
                missing_counter[rel] += 1

        ws.cell(r, col_present_total).value = f"{present}/{len(refs)}"
        ws.cell(r, col_missing).value = "\n".join(missing)
        ws.cell(r, col_sha).value = "\n".join(sha_entries)
        ws.cell(r, col_evid).value = "\n".join(sorted(ev_ids))

        if present == len(refs):
            comp = "ALL_PRESENT"
        elif present == 0:
            comp = "ALL_MISSING"
        else:
            comp = "PARTIAL"
        ws.cell(r, col_comp).value = comp

    # Summary sheet
    if "Autofill_Summary" in wb.sheetnames:
        del wb["Autofill_Summary"]
    ws_sum = wb.create_sheet("Autofill_Summary")
    ws_sum["A1"].value = "Build root"
    ws_sum["B1"].value = str(build_root)
    ws_sum["A2"].value = "GSPR input"
    ws_sum["B2"].value = str(gspr_in)
    ws_sum["A3"].value = "Evidence manifest"
    ws_sum["B3"].value = str(evidence_manifest) if evidence_manifest else "(none)"

    ws_sum["A5"].value = "Total referenced files"
    ws_sum["B5"].value = total_refs
    ws_sum["A6"].value = "Total present"
    ws_sum["B6"].value = total_present
    ws_sum["A7"].value = "Total missing"
    ws_sum["B7"].value = max(0, total_refs - total_present)

    ws_sum["A9"].value = "Missing reference"
    ws_sum["B9"].value = "Count"

    row = 10
    for rel, cnt in missing_counter.most_common():
        ws_sum.cell(row, 1).value = rel
        ws_sum.cell(row, 2).value = cnt
        row += 1

    gspr_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(gspr_out)
    print(f"[OK] Wrote: {gspr_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
