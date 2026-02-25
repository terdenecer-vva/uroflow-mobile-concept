#!/usr/bin/env python3
"""build_dataset_release_bundle_guarded.py  (v4.2)

Creates a DatasetRelease bundle ONLY if:
1) Pre-Freeze Gates PASS (dataset-level readiness), AND
2) Record-level gates are computed, AND
3) The release contains at least N valid records (default N=1).

Key upgrade vs v3.8:
- Record-level filtering: the release includes *only* records that pass required per-record gates.
- The release contains:
  - manifest_included.csv
  - manifest_excluded.csv
  - record_level_gates.csv (+ summary)

This design prevents a common failure mode:
"We froze a dataset_id, but later discovered that 20% of records were invalid due to privacy/sync/stand pose."

Inputs:
- --dataset_root : dataset root containing records/ and outputs/
- --manifest     : CSV/XLSX manifest
- --dataset_id   : optional; if omitted -> auto generated
- --pre_freeze_report : optional; default outputs/pre_freeze_gates/pre_freeze_gates_report.json
- --min_included_records : optional; default 1

Outputs (under dataset_root):
- outputs/dataset_release/<dataset_id>/
- outputs/dataset_release/dataset_release_<dataset_id>.zip
- outputs/dataset_release/release_blocked.json (if blocked)

Also logs freeze event to a DHF-friendly Freeze Event Log (XLSX) under the Submission Build tree.

Note:
- This script may auto-run validators if their outputs are missing:
  validate_privacy_live_guardrails.py
  validate_ios_capture_contract.py
  validate_privacy_guardrails_consistency.py
  compute_record_level_gates.py

It does NOT auto-run heavy video processing by default.
If ROI video is stored and content-level checks are required, run validate_privacy_content_guardrails_v2.py beforehand.
"""

from __future__ import annotations

import argparse
import json
import shutil
import zipfile
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List
import hashlib
import re
import subprocess
import sys

import pandas as pd
import openpyxl


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def derive_lock_id_from_filename(fn: str) -> str:
    """Derive a stable lock identifier from a lock filename.

    Example:
      Uroflow_IntendedUse_Claims_ByRegion_Lock_v2.0_EN.docx -> LOCK_v2.0
    """
    if "Lock_v" in fn:
        v = fn.split("Lock_v", 1)[1].split("_", 1)[0].replace(".docx", "")
        return f"LOCK_v{v}"
    return "LOCK_UNSPEC"


def pick_latest_lock_file(folder: Path, pattern_list: list[str]) -> Path | None:
    candidates: list[Path] = []
    for pat in pattern_list:
        candidates.extend(folder.glob(pat))
    if not candidates:
        return None

    def parse_ver(p: Path) -> tuple:
        # best-effort parse 'vX.Y' from filename
        m = re.search(r"_v(\d+(?:\.\d+)?)", p.name)
        if not m:
            return (0.0, p.name)
        try:
            return (float(m.group(1)), p.name)
        except Exception:
            return (0.0, p.name)

    candidates.sort(key=parse_ver, reverse=True)
    return candidates[0]



def load_manifest(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    if "record_id" not in df.columns:
        raise ValueError("Manifest must contain record_id.")
    df["record_id"] = df["record_id"].astype(str)
    return df


def ensure_freeze_log(build_root: Path, freeze_log_xlsx: Path | None) -> Path:
    if freeze_log_xlsx is not None:
        freeze_log_xlsx.parent.mkdir(parents=True, exist_ok=True)
        if not freeze_log_xlsx.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "FreezeEvents"
            ws.append([
                "event_id","timestamp_utc","operator_id","event_type","dataset_id","model_id",
                "claims_lock_id","acceptance_lock_id","pre_freeze_report_path","pre_freeze_report_sha256",
                "release_bundle_path","release_bundle_sha256","notes"
            ])
            ws.freeze_panes = "A2"
            wb.save(freeze_log_xlsx)
        return freeze_log_xlsx

    template = build_root / "15_Dataset_Model_Release" / "Uroflow_DHF_Freeze_Event_Log_Template_v1.0.xlsx"
    freeze_log = build_root / "15_Dataset_Model_Release" / "DHF_Freeze_Event_Log.xlsx"
    freeze_log.parent.mkdir(parents=True, exist_ok=True)
    if freeze_log.exists():
        return freeze_log
    if template.exists():
        shutil.copyfile(template, freeze_log)
        return freeze_log

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FreezeEvents"
    ws.append([
        "event_id","timestamp_utc","operator_id","event_type","dataset_id","model_id",
        "claims_lock_id","acceptance_lock_id","pre_freeze_report_path","pre_freeze_report_sha256",
        "release_bundle_path","release_bundle_sha256","notes"
    ])
    ws.freeze_panes = "A2"
    wb.save(freeze_log)
    return freeze_log


def append_freeze_event(freeze_log: Path, row: List[Any]) -> None:
    wb = openpyxl.load_workbook(freeze_log)
    ws = wb["FreezeEvents"] if "FreezeEvents" in wb.sheetnames else wb.active
    ws.append(row)
    wb.save(freeze_log)


def run_if_missing(out_path: Path, cmd: List[str], cwd: Path) -> None:
    if out_path.exists():
        return
    print(f"[AUTO] Missing {out_path.name} -> running: {' '.join(cmd)}")
    res = subprocess.run(cmd, cwd=str(cwd))
    if res.returncode != 0:
        raise RuntimeError(f"Command failed ({res.returncode}): {' '.join(cmd)}")
    if not out_path.exists():
        raise RuntimeError(f"Expected output not produced: {out_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset_root", required=True)
    ap.add_argument("--manifest", required=True)
    ap.add_argument("--dataset_id", default=None)
    ap.add_argument("--pre_freeze_report", default=None)
    ap.add_argument("--min_included_records", type=int, default=1)

    ap.add_argument("--operator_id", default="UNKNOWN")
    ap.add_argument("--claims_lock_id", default="")
    ap.add_argument("--acceptance_lock_id", default="")
    ap.add_argument("--notes", default="")
    ap.add_argument("--freeze_log_xlsx", default=None)
    args = ap.parse_args()

    dataset_root = Path(args.dataset_root)
    manifest = Path(args.manifest)

    # Identify submission build root based on script location
    build_root = Path(__file__).resolve().parents[2]
    auto_dir = Path(__file__).resolve().parents[1]  # 10_Pilot_Automation


    # Auto-detect lock files/ids if not provided (keeps releases traceable by default)
    qms_dir = build_root / "01_Product_QMS"
    claims_lock_file = pick_latest_lock_file(qms_dir, [
        "Uroflow_IntendedUse_Claims_ByRegion_Lock_v*_EN.docx",
        "Uroflow_IntendedUse_Claims_ByRegion_Lock_v*_RU.docx",
    ])
    acceptance_lock_file = pick_latest_lock_file(qms_dir, [
        "Uroflow_Acceptance_Criteria_Lock_v*_EN.docx",
        "Uroflow_Acceptance_Criteria_Lock_v*_RU.docx",
    ])
    if not args.claims_lock_id and claims_lock_file:
        args.claims_lock_id = derive_lock_id_from_filename(claims_lock_file.name)
    if not args.acceptance_lock_id and acceptance_lock_file:
        args.acceptance_lock_id = derive_lock_id_from_filename(acceptance_lock_file.name)


    pre_freeze_report = Path(args.pre_freeze_report) if args.pre_freeze_report else (dataset_root / "outputs/pre_freeze_gates/pre_freeze_gates_report.json")
    if not pre_freeze_report.exists():
        raise FileNotFoundError(f"Pre-freeze report not found: {pre_freeze_report}")

    rep = json.loads(pre_freeze_report.read_text(encoding="utf-8"))
    if not bool(rep.get("overall_pass", False)):
        out_dir = dataset_root / "outputs/dataset_release"
        out_dir.mkdir(parents=True, exist_ok=True)
        blocked = {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "status": "NO_RELEASE",
            "reason": "PRE_FREEZE_GATES_FAILED",
            "required_failed": rep.get("required_failed", []),
            "pre_freeze_report": str(pre_freeze_report),
        }
        (out_dir / "release_blocked.json").write_text(json.dumps(blocked, indent=2), encoding="utf-8")
        print("[NO_RELEASE] Pre-freeze gates failed. See outputs/dataset_release/release_blocked.json")
        raise SystemExit(2)

    # Ensure prerequisite validators exist (lightweight)
    # 1) privacy live csv
    live_csv = dataset_root / "outputs/privacy_live_guardrails/privacy_live_guardrails.csv"
    run_if_missing(
        live_csv,
        [sys.executable, "scripts/validate_privacy_live_guardrails.py", "--dataset_root", str(dataset_root), "--manifest", str(manifest)],
        cwd=auto_dir
    )

    # 2) iOS contract validation
    ios_csv = dataset_root / "outputs/ios_capture_contract/ios_capture_contract_validation.csv"
    run_if_missing(
        ios_csv,
        [sys.executable, "scripts/validate_ios_capture_contract.py", "--dataset_root", str(dataset_root), "--manifest", str(manifest)],
        cwd=auto_dir
    )

    # 3) privacy consistency report
    cons_csv = dataset_root / "outputs/privacy_consistency/privacy_consistency.csv"
    run_if_missing(
        cons_csv,
        [sys.executable, "scripts/validate_privacy_guardrails_consistency.py", "--dataset_root", str(dataset_root), "--manifest", str(manifest)],
        cwd=auto_dir
    )

    # 4) record-level gates
    gates_csv = dataset_root / "outputs/record_level_gates/record_level_gates.csv"
    run_if_missing(
        gates_csv,
        [sys.executable, "scripts/compute_record_level_gates.py", "--dataset_root", str(dataset_root), "--manifest", str(manifest)],
        cwd=auto_dir
    )

    gates_df = pd.read_csv(gates_csv)
    gates_df["record_id"] = gates_df["record_id"].astype(str)
    included_ids = gates_df[gates_df["include_in_release"] == True]["record_id"].astype(str).tolist()  # noqa: E712
    excluded_ids = gates_df[gates_df["include_in_release"] != True]["record_id"].astype(str).tolist()  # noqa: E712

    if len(included_ids) < int(args.min_included_records):
        out_dir = dataset_root / "outputs/dataset_release"
        out_dir.mkdir(parents=True, exist_ok=True)
        blocked = {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "status": "NO_RELEASE",
            "reason": "INSUFFICIENT_VALID_RECORDS",
            "min_required": int(args.min_included_records),
            "included": int(len(included_ids)),
            "excluded": int(len(excluded_ids)),
            "record_level_gates_csv": str(gates_csv),
        }
        (out_dir / "release_blocked.json").write_text(json.dumps(blocked, indent=2), encoding="utf-8")
        print("[NO_RELEASE] Too few valid records. See outputs/dataset_release/release_blocked.json")
        raise SystemExit(3)

    # generate dataset_id if needed
    dataset_id = args.dataset_id or ("UF-GD-" + datetime.utcnow().strftime("%Y%m%d-%H%M%S"))
    release_dir = dataset_root / "outputs/dataset_release" / dataset_id
    release_dir.mkdir(parents=True, exist_ok=True)

    # copy original inputs
    shutil.copyfile(manifest, release_dir / ("manifest_original" + manifest.suffix.lower()))
    shutil.copyfile(pre_freeze_report, release_dir / "pre_freeze_gates_report.json")
    pre_sum = pre_freeze_report.parent / "pre_freeze_gates_summary.txt"
    if pre_sum.exists():
        shutil.copyfile(pre_sum, release_dir / "pre_freeze_gates_summary.txt")

    # write filtered manifests
    dfm = load_manifest(manifest)
    df_in = dfm[dfm["record_id"].isin(included_ids)].copy()
    df_ex = dfm[~dfm["record_id"].isin(included_ids)].copy()
    df_in.to_csv(release_dir / "manifest_included.csv", index=False)
    df_ex.to_csv(release_dir / "manifest_excluded.csv", index=False)

    # copy record-level gates
    shutil.copyfile(gates_csv, release_dir / "record_level_gates.csv")
    gates_sum = dataset_root / "outputs/record_level_gates/record_level_gates_summary.json"
    if gates_sum.exists():
        shutil.copyfile(gates_sum, release_dir / "record_level_gates_summary.json")

    included_files = [
        release_dir.name + "/" + ("manifest_original" + manifest.suffix.lower()),
        release_dir.name + "/pre_freeze_gates_report.json",
        release_dir.name + "/manifest_included.csv",
        release_dir.name + "/manifest_excluded.csv",
        release_dir.name + "/record_level_gates.csv",
    ]
    if (release_dir / "pre_freeze_gates_summary.txt").exists():
        included_files.append(release_dir.name + "/pre_freeze_gates_summary.txt")
    if (release_dir / "record_level_gates_summary.json").exists():
        included_files.append(release_dir.name + "/record_level_gates_summary.json")

    rel_manifest = {
        "dataset_id": dataset_id,
        "created_at": datetime.utcnow().isoformat() + "Z",
        "dataset_root": str(dataset_root),
        "source_manifest": str(manifest),
        "pre_freeze_report": str(pre_freeze_report),
        "record_level_gates_csv": str(gates_csv),
        "included_records": int(len(included_ids)),
        "excluded_records": int(len(excluded_ids)),
        "claims_lock_id": args.claims_lock_id,
        "claims_lock_file": str(claims_lock_file) if "claims_lock_file" in locals() and claims_lock_file else "",
        "claims_lock_file_sha256": sha256_file(claims_lock_file) if "claims_lock_file" in locals() and claims_lock_file and claims_lock_file.exists() else "",
        "acceptance_lock_id": args.acceptance_lock_id,
        "acceptance_lock_file": str(acceptance_lock_file) if "acceptance_lock_file" in locals() and acceptance_lock_file else "",
        "acceptance_lock_file_sha256": sha256_file(acceptance_lock_file) if "acceptance_lock_file" in locals() and acceptance_lock_file and acceptance_lock_file.exists() else "",
        "operator_id": args.operator_id,
        "notes": args.notes,
    }
    (release_dir / "dataset_release_manifest.json").write_text(json.dumps(rel_manifest, indent=2), encoding="utf-8")

    # checksums (files inside release_dir only)
    checksum_lines = []
    for fn in ["manifest_original" + manifest.suffix.lower(), "pre_freeze_gates_report.json", "manifest_included.csv", "manifest_excluded.csv", "record_level_gates.csv", "dataset_release_manifest.json"]:
        p = release_dir / fn
        checksum_lines.append(f"{sha256_file(p)}  {fn}")
    if (release_dir / "pre_freeze_gates_summary.txt").exists():
        checksum_lines.append(f"{sha256_file(release_dir / 'pre_freeze_gates_summary.txt')}  pre_freeze_gates_summary.txt")
    if (release_dir / "record_level_gates_summary.json").exists():
        checksum_lines.append(f"{sha256_file(release_dir / 'record_level_gates_summary.json')}  record_level_gates_summary.json")
    (release_dir / "checksums.sha256").write_text("\n".join(checksum_lines) + "\n", encoding="utf-8")

    # zip bundle
    zip_path = dataset_root / "outputs/dataset_release" / f"dataset_release_{dataset_id}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for fn in ["manifest_original" + manifest.suffix.lower(), "pre_freeze_gates_report.json", "manifest_included.csv", "manifest_excluded.csv", "record_level_gates.csv", "dataset_release_manifest.json", "checksums.sha256"]:
            z.write(release_dir / fn, arcname=f"{dataset_id}/{fn}")
        if (release_dir / "pre_freeze_gates_summary.txt").exists():
            z.write(release_dir / "pre_freeze_gates_summary.txt", arcname=f"{dataset_id}/pre_freeze_gates_summary.txt")
        if (release_dir / "record_level_gates_summary.json").exists():
            z.write(release_dir / "record_level_gates_summary.json", arcname=f"{dataset_id}/record_level_gates_summary.json")

    # log freeze event to DHF
    freeze_log_path = Path(args.freeze_log_xlsx) if args.freeze_log_xlsx else None
    freeze_log = ensure_freeze_log(build_root, freeze_log_path)

    event_id = f"EV-FREEZE-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}"
    note = (args.notes + " | " if args.notes else "") + f"included={len(included_ids)} excluded={len(excluded_ids)}"
    row = [
        event_id,
        datetime.utcnow().isoformat() + "Z",
        args.operator_id,
        "DatasetRelease",
        dataset_id,
        "",
        args.claims_lock_id,
        args.acceptance_lock_id,
        str(pre_freeze_report),
        sha256_file(pre_freeze_report),
        str(zip_path),
        sha256_file(zip_path),
        note,
    ]
    append_freeze_event(freeze_log, row)

    # write receipt under dataset_root
    receipt = {
        "event_id": event_id,
        "timestamp_utc": row[1],
        "dataset_id": dataset_id,
        "included_records": int(len(included_ids)),
        "excluded_records": int(len(excluded_ids)),
        "release_zip": str(zip_path),
        "freeze_log": str(freeze_log),
    }
    (dataset_root / "outputs/freeze_events").mkdir(parents=True, exist_ok=True)
    (dataset_root / "outputs/freeze_events" / f"freeze_receipt_{dataset_id}.json").write_text(json.dumps(receipt, indent=2), encoding="utf-8")



    # Create Freeze Kit automatically (mandatory for pilot-ready execution).
    try:
        freeze_kit_script = auto_dir / "scripts" / "build_pilot_freeze_kit.py"
        if freeze_kit_script.exists():
            subprocess.run(
                [
                    sys.executable,
                    str(freeze_kit_script),
                    "--dataset_root", str(dataset_root),
                    "--dataset_id", dataset_id,
                    "--operator_id", str(args.operator_id),
                ],
                cwd=str(auto_dir),
                check=True,
            )
        else:
            print("[WARN] Freeze Kit script not found; skipping Freeze Kit generation.")
    except Exception as exc:  # pragma: no cover - best-effort follow-up step
        print(f"[WARN] Freeze Kit generation failed: {exc}")

    print(f"[OK] DatasetRelease created: {zip_path}")
    print(f"[OK] Included records: {len(included_ids)} | Excluded: {len(excluded_ids)}")
    print(f"[OK] Freeze event logged: {freeze_log} (event_id={event_id})")


if __name__ == "__main__":
    main()
