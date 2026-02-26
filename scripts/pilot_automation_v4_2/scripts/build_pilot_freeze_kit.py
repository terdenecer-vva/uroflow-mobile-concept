#!/usr/bin/env python3
"""build_pilot_freeze_kit.py  (v4.2)

Creates an executed Freeze Kit artifact binding:
- dataset_id
- DatasetRelease ZIP
- Claims Lock (file + SHA256)
- Acceptance Lock (file + SHA256)
- Key gate outputs (paths + SHA256)

Outputs (under Submission Build tree):
- 15_Dataset_Model_Release/Freeze_Kits/FreezeKit_<dataset_id>_<timestampUTC>.xlsx
- 15_Dataset_Model_Release/Freeze_Kits/FreezeKit_<dataset_id>_<timestampUTC>.json
- 15_Dataset_Model_Release/Freeze_Kits/FreezeKit_<dataset_id>_<timestampUTC>_checksums.sha256

Also appends a row to DHF_Freeze_Event_Log.xlsx (event_type=FreezeKit).
"""

from __future__ import annotations
import argparse
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List

import openpyxl
import pandas as pd


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def pick_existing(paths: List[Path]) -> Optional[Path]:
    for p in paths:
        if p.exists():
            return p
    return None


def ensure_freeze_log(build_root: Path) -> Path:
    template = build_root / "15_Dataset_Model_Release" / "Uroflow_DHF_Freeze_Event_Log_Template_v1.0.xlsx"
    freeze_log = build_root / "15_Dataset_Model_Release" / "DHF_Freeze_Event_Log.xlsx"
    freeze_log.parent.mkdir(parents=True, exist_ok=True)
    if freeze_log.exists():
        return freeze_log
    if template.exists():
        shutil.copyfile(template, freeze_log)
        return freeze_log

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FreezeEvents"
    ws.append([
        "event_id","timestamp_utc","operator_id","event_type","dataset_id","model_id",
        "claims_lock_id","acceptance_lock_id","pre_freeze_report_path","pre_freeze_report_sha256",
        "release_bundle_path","release_bundle_sha256","notes"
    ])
    ws.freeze_panes = "A2"
    wb.save(freeze_log)
    return freeze_log


def append_freeze_event(freeze_log: Path, row: List[Any]) -> None:
    wb = openpyxl.load_workbook(freeze_log)
    ws = wb["FreezeEvents"] if "FreezeEvents" in wb.sheetnames else wb.active
    ws.append(row)
    wb.save(freeze_log)


def derive_lock_id_from_filename(fn: str) -> str:
    # Example: Uroflow_IntendedUse_Claims_ByRegion_Lock_v2.0_EN.docx -> CLAIMS_LOCK_v2.0
    if "Lock_v" in fn:
        v = fn.split("Lock_v", 1)[1].split("_", 1)[0].replace(".docx", "")
        return f"LOCK_v{v}"
    return "LOCK_UNSPEC"


def set_value_by_label(ws, label: str, value: Any) -> None:
    # Search column A for label and set column B
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value).strip() == label:
            ws.cell(row=row, column=2).value = value
            return


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset_root", required=True)
    ap.add_argument("--dataset_id", default=None)
    ap.add_argument("--operator_id", default="UNKNOWN")
    ap.add_argument("--release_zip", default=None)
    ap.add_argument("--pre_freeze_report", default=None)
    ap.add_argument("--claims_lock_file", default=None)
    ap.add_argument("--acceptance_lock_file", default=None)
    ap.add_argument("--template_xlsx", default=None)
    ap.add_argument("--output_dir", default=None)
    args = ap.parse_args()

    dataset_root = Path(args.dataset_root)
    if not dataset_root.exists():
        raise FileNotFoundError(f"dataset_root not found: {dataset_root}")

    # Locate Submission Build root (two levels above scripts/)
    build_root = Path(__file__).resolve().parents[2]

    # Determine dataset_id
    if args.dataset_id:
        dataset_id = args.dataset_id
    else:
        rel_dir = dataset_root / "outputs/dataset_release"
        if not rel_dir.exists():
            raise FileNotFoundError("No outputs/dataset_release directory; provide --dataset_id and --release_zip.")
        # pick the newest subdir that looks like a dataset_id directory
        candidates = [p for p in rel_dir.iterdir() if p.is_dir()]
        if not candidates:
            raise FileNotFoundError("No dataset release directories found under outputs/dataset_release")
        candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        dataset_id = candidates[0].name

    release_zip = Path(args.release_zip) if args.release_zip else (dataset_root / "outputs/dataset_release" / f"dataset_release_{dataset_id}.zip")
    if not release_zip.exists():
        raise FileNotFoundError(f"DatasetRelease ZIP not found: {release_zip}")

    pre_freeze_report = Path(args.pre_freeze_report) if args.pre_freeze_report else (dataset_root / "outputs/pre_freeze_gates/pre_freeze_gates_report.json")
    if not pre_freeze_report.exists():
        raise FileNotFoundError(f"Pre-freeze report not found: {pre_freeze_report}")

    # Default lock files
    default_claims = pick_existing([
        build_root / "01_Product_QMS" / "Uroflow_IntendedUse_Claims_ByRegion_Lock_v2.0_EN.docx",
        build_root / "01_Product_QMS" / "Uroflow_IntendedUse_Claims_ByRegion_Lock_v2.0_RU.docx",
    ])
    default_acceptance = pick_existing([
        build_root / "01_Product_QMS" / "Uroflow_Acceptance_Criteria_Lock_v2.0_EN.docx",
        build_root / "01_Product_QMS" / "Uroflow_Acceptance_Criteria_Lock_v2.0_RU.docx",
    ])
    claims_lock_file = Path(args.claims_lock_file) if args.claims_lock_file else default_claims
    acceptance_lock_file = Path(args.acceptance_lock_file) if args.acceptance_lock_file else default_acceptance
    if claims_lock_file is None or not claims_lock_file.exists():
        raise FileNotFoundError("Claims lock file not found. Provide --claims_lock_file.")
    if acceptance_lock_file is None or not acceptance_lock_file.exists():
        raise FileNotFoundError("Acceptance lock file not found. Provide --acceptance_lock_file.")

    claims_lock_id = derive_lock_id_from_filename(claims_lock_file.name)
    acceptance_lock_id = derive_lock_id_from_filename(acceptance_lock_file.name)

    # Site IDs summary (best-effort)
    site_summary = ""
    release_dir = dataset_root / "outputs/dataset_release" / dataset_id
    manifest_included = release_dir / "manifest_included.csv"
    if manifest_included.exists():
        try:
            df = pd.read_csv(manifest_included)
            if "site_id" in df.columns:
                vc = df["site_id"].astype(str).value_counts()
                site_summary = ", ".join([f"{k}({int(v)})" for k,v in vc.items()])
        except Exception:
            site_summary = ""

    # Locate other outputs (best-effort)
    def out(p: str) -> Path:
        return dataset_root / p

    record_level_sum = out("outputs/record_level_gates/record_level_gates_summary.json")
    record_level_csv = out("outputs/record_level_gates/record_level_gates.csv")
    coverage_sum = out("outputs/coverage_dashboard/coverage_summary.json")
    acc_sum = out("outputs/accuracy_acceptance/accuracy_summary.json")

    # Load template
    template_xlsx = Path(args.template_xlsx) if args.template_xlsx else (build_root / "15_Dataset_Model_Release" / "Uroflow_Pilot_Freeze_Kit_Template_v1.0.xlsx")
    if not template_xlsx.exists():
        raise FileNotFoundError(f"Freeze Kit template not found: {template_xlsx}")
    wb = openpyxl.load_workbook(template_xlsx)
    ws = wb["Freeze_Kit"] if "Freeze_Kit" in wb.sheetnames else wb.active

    now = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    created_at = datetime.utcnow().isoformat() + "Z"

    set_value_by_label(ws, "Dataset ID", dataset_id)
    set_value_by_label(ws, "Created at (UTC)", created_at)
    set_value_by_label(ws, "Operator ID", args.operator_id)
    set_value_by_label(ws, "Site IDs (summary)", site_summary)
    set_value_by_label(ws, "DatasetRelease ZIP path", str(release_zip))
    set_value_by_label(ws, "DatasetRelease ZIP SHA256", sha256_file(release_zip))

    set_value_by_label(ws, "Claims Lock ID", claims_lock_id)
    set_value_by_label(ws, "Claims Lock file path", str(claims_lock_file))
    set_value_by_label(ws, "Claims Lock SHA256", sha256_file(claims_lock_file))

    set_value_by_label(ws, "Acceptance Lock ID", acceptance_lock_id)
    set_value_by_label(ws, "Acceptance Lock file path", str(acceptance_lock_file))
    set_value_by_label(ws, "Acceptance Lock SHA256", sha256_file(acceptance_lock_file))

    set_value_by_label(ws, "Pre-freeze gates report path", str(pre_freeze_report))
    set_value_by_label(ws, "Pre-freeze gates report SHA256", sha256_file(pre_freeze_report))

    # record-level gates summary preference
    if record_level_sum.exists():
        set_value_by_label(ws, "Record-level gates summary path", str(record_level_sum))
        set_value_by_label(ws, "Record-level gates summary SHA256", sha256_file(record_level_sum))
    elif record_level_csv.exists():
        set_value_by_label(ws, "Record-level gates summary path", str(record_level_csv))
        set_value_by_label(ws, "Record-level gates summary SHA256", sha256_file(record_level_csv))

    if coverage_sum.exists():
        set_value_by_label(ws, "Coverage summary path", str(coverage_sum))
        set_value_by_label(ws, "Coverage summary SHA256", sha256_file(coverage_sum))

    if acc_sum.exists():
        set_value_by_label(ws, "Acceptance metrics summary path (optional)", str(acc_sum))
        set_value_by_label(ws, "Acceptance metrics summary SHA256 (optional)", sha256_file(acc_sum))

    # Fill artifact table (best-effort): find header row where col1 == 'Artifact'
    header_row = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r,1).value).strip() == "Artifact" and str(ws.cell(r,2).value).strip() == "Required":
            header_row = r
            break
    if header_row:
        r = header_row + 1
        while r <= ws.max_row and ws.cell(r,1).value:
            art = str(ws.cell(r,1).value).strip()
            # map artifact name to known output paths
            mapping = {
                "pre_freeze_gates_report.json": pre_freeze_report,
                "record_level_gates_summary.json": record_level_sum if record_level_sum.exists() else record_level_csv,
                "coverage_summary.json": coverage_sum,
                "privacy_live_guardrails.csv": out("outputs/privacy_live_guardrails/privacy_live_guardrails.csv"),
                "ios_capture_contract_validation.csv": out("outputs/ios_capture_contract/ios_capture_contract_validation.csv"),
                "privacy_consistency.csv": out("outputs/privacy_consistency/privacy_consistency.csv"),
                "stand_pose_drift_summary.json": out("outputs/stand_pose_drift_dashboard/drift_summary.json"),
                "privacy_content_guardrails_v2_summary.json": out("outputs/privacy_content_guardrails_v2/privacy_content_guardrails_v2_summary.json"),
                "accuracy_summary.json": acc_sum,
            }
            p = mapping.get(art)
            if p and isinstance(p, Path) and p.exists():
                ws.cell(r,3).value = str(p.relative_to(dataset_root)) if str(p).startswith(str(dataset_root)) else str(p)
                ws.cell(r,4).value = sha256_file(p)
            r += 1

    out_dir = Path(args.output_dir) if args.output_dir else (build_root / "15_Dataset_Model_Release" / "Freeze_Kits")
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_out = out_dir / f"FreezeKit_{dataset_id}_{now}Z.xlsx"
    wb.save(xlsx_out)

    summary: Dict[str, Any] = {
        "freeze_kit_version": "v1.0",
        "dataset_id": dataset_id,
        "created_at_utc": created_at,
        "operator_id": args.operator_id,
        "dataset_release_zip": str(release_zip),
        "dataset_release_zip_sha256": sha256_file(release_zip),
        "claims_lock_id": claims_lock_id,
        "claims_lock_file": str(claims_lock_file),
        "claims_lock_sha256": sha256_file(claims_lock_file),
        "acceptance_lock_id": acceptance_lock_id,
        "acceptance_lock_file": str(acceptance_lock_file),
        "acceptance_lock_sha256": sha256_file(acceptance_lock_file),
        "pre_freeze_report": str(pre_freeze_report),
        "pre_freeze_report_sha256": sha256_file(pre_freeze_report),
        "freeze_kit_xlsx": str(xlsx_out),
        "freeze_kit_xlsx_sha256": sha256_file(xlsx_out),
    }
    json_out = out_dir / f"FreezeKit_{dataset_id}_{now}Z.json"
    json_out.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    checksums = out_dir / f"FreezeKit_{dataset_id}_{now}Z_checksums.sha256"
    checksums.write_text(
        f"{sha256_file(xlsx_out)}  {xlsx_out.name}\n" +
        f"{sha256_file(json_out)}  {json_out.name}\n",
        encoding="utf-8"
    )

    # Append to freeze log
    freeze_log = ensure_freeze_log(build_root)
    event_id = f"EV-FREEZE-KIT-{now}"
    notes = f"FreezeKit created; DatasetRelease={release_zip.name}"
    row = [
        event_id,
        created_at,
        args.operator_id,
        "FreezeKit",
        dataset_id,
        "",
        claims_lock_id,
        acceptance_lock_id,
        str(pre_freeze_report),
        sha256_file(pre_freeze_report),
        str(xlsx_out),
        sha256_file(xlsx_out),
        notes,
    ]
    append_freeze_event(freeze_log, row)

    print(f"[OK] Freeze Kit created: {xlsx_out}")
    print(f"[OK] Freeze Kit summary: {json_out}")
    print(f"[OK] Freeze event appended: {freeze_log} (event_id={event_id})")


if __name__ == "__main__":
    main()
