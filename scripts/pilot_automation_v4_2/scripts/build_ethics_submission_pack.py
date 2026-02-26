#!/usr/bin/env python3
"""build_ethics_submission_pack.py  (v4.2)

Builds a region-specific Ethics/IRB submission ZIP based on:
- Submission_Build_v4.2/19_Ethics_Submission_Packs/Ethics_Submission_Pack_Index_v1.0.xlsx

It copies files with Include_in_pack == 'Y' into an output folder and generates:
- pack_manifest.json
- checksums.sha256
- ethics_pack_<REGION>_<timestampUTC>.zip

Usage:
  python build_ethics_submission_pack.py --region RU_EC
"""

from __future__ import annotations
import argparse
import hashlib
import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

import openpyxl


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def zip_dir(src_dir: Path, zip_path: Path) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for p in src_dir.rglob("*"):
            if p.is_file():
                z.write(p, arcname=str(p.relative_to(src_dir)))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--region", required=True, choices=["RU_EC","EU_Ethics","US_IRB"])
    ap.add_argument("--index_xlsx", default=None)
    ap.add_argument("--output_base", default=None)
    args = ap.parse_args()

    build_root = Path(__file__).resolve().parents[2]
    index_xlsx = Path(args.index_xlsx) if args.index_xlsx else (build_root / "19_Ethics_Submission_Packs" / "Ethics_Submission_Pack_Index_v1.0.xlsx")
    if not index_xlsx.exists():
        raise FileNotFoundError(f"Index not found: {index_xlsx}")

    wb = openpyxl.load_workbook(index_xlsx)
    if args.region not in wb.sheetnames:
        raise ValueError(f"Sheet not found in index: {args.region}")
    ws = wb[args.region]

    now = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    out_base = Path(args.output_base) if args.output_base else (build_root / "19_Ethics_Submission_Packs" / "Output_Packs")
    out_dir = out_base / f"{args.region}_{now}Z"
    out_dir.mkdir(parents=True, exist_ok=True)

    # header mapping from row1
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    def col_idx(name: str) -> int:
        if name not in header:
            raise ValueError(f"Column '{name}' not found in sheet {args.region}")
        return header.index(name) + 1

    idx_include = col_idx("Include_in_pack")
    idx_path = col_idx("Build relative path")
    idx_docid = col_idx("Doc_ID")
    idx_docname = col_idx("Document name")

    included: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        inc = ws.cell(r, idx_include).value
        rel = ws.cell(r, idx_path).value
        if not rel:
            continue
        if str(inc).strip().upper() != "Y":
            continue
        src = build_root / str(rel).strip()
        if not src.exists():
            raise FileNotFoundError(f"Missing file listed in index: {rel}")
        dst = out_dir / str(rel).strip()
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(src, dst)
        included.append({
            "doc_id": ws.cell(r, idx_docid).value,
            "doc_name": ws.cell(r, idx_docname).value,
            "relative_path": str(rel).strip(),
            "sha256": sha256_file(dst),
            "size_bytes": dst.stat().st_size,
        })

    manifest = {
        "region": args.region,
        "generated_at_utc": datetime.utcnow().isoformat() + "Z",
        "source_index": str(index_xlsx),
        "output_dir": str(out_dir),
        "files": included,
    }
    (out_dir / "pack_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    # checksums
    lines = []
    for f in sorted([p for p in out_dir.rglob("*") if p.is_file() and p.name != "checksums.sha256"], key=lambda p: str(p)):
        lines.append(f"{sha256_file(f)}  {f.relative_to(out_dir)}")
    (out_dir / "checksums.sha256").write_text("\n".join(lines) + "\n", encoding="utf-8")

    zip_path = out_base / f"ethics_pack_{args.region}_{now}Z.zip"
    zip_dir(out_dir, zip_path)
    print(f"[OK] Ethics submission pack ZIP: {zip_path}")


if __name__ == "__main__":
    main()
