#!/usr/bin/env python3
"""
log_freeze_event_to_dhf.py

Appends a freeze event (DatasetRelease or ModelRelease) to a DHF-friendly Freeze Event Log (XLSX).

This is a helper utility used by guarded builders.
"""
from __future__ import annotations

import argparse
from pathlib import Path
from datetime import datetime
import hashlib
import openpyxl


def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--freeze_log_xlsx", required=True)
    ap.add_argument("--event_type", required=True, choices=["DatasetRelease","ModelRelease"])
    ap.add_argument("--operator_id", default="UNKNOWN")
    ap.add_argument("--dataset_id", default="")
    ap.add_argument("--model_id", default="")
    ap.add_argument("--claims_lock_id", default="")
    ap.add_argument("--acceptance_lock_id", default="")
    ap.add_argument("--pre_freeze_report", required=True)
    ap.add_argument("--release_bundle", required=True)
    ap.add_argument("--notes", default="")
    args = ap.parse_args()

    log_path = Path(args.freeze_log_xlsx)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    if not log_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FreezeEvents"
        ws.append([
            "event_id","timestamp_utc","operator_id","event_type","dataset_id","model_id",
            "claims_lock_id","acceptance_lock_id","pre_freeze_report_path","pre_freeze_report_sha256",
            "release_bundle_path","release_bundle_sha256","notes"
        ])
        ws.freeze_panes = "A2"
        wb.save(log_path)

    wb = openpyxl.load_workbook(log_path)
    ws = wb["FreezeEvents"] if "FreezeEvents" in wb.sheetnames else wb.active

    event_id = f"EV-FREEZE-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}"
    row = [
        event_id,
        datetime.utcnow().isoformat() + "Z",
        args.operator_id,
        args.event_type,
        args.dataset_id,
        args.model_id,
        args.claims_lock_id,
        args.acceptance_lock_id,
        str(Path(args.pre_freeze_report)),
        sha256_file(Path(args.pre_freeze_report)),
        str(Path(args.release_bundle)),
        sha256_file(Path(args.release_bundle)),
        args.notes,
    ]
    ws.append(row)
    wb.save(log_path)
    print(f"[OK] Logged {args.event_type} event_id={event_id} -> {log_path}")


if __name__ == "__main__":
    main()
